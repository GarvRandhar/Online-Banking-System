from email import errors
import os
import re
import string
import random
import csv
import logging
import smtplib
import tempfile
import openpyxl
import calendar
import secrets
from datetime import datetime, timedelta
import pyotp
import qrcode
from io import BytesIO, StringIO
import base64
from sqlalchemy import inspect
from flask import abort, Flask, render_template, request, redirect, url_for, flash, session
from flask_wtf.csrf import CSRFProtect, CSRFError, generate_csrf, validate_csrf
from models import db, Account, Amount, Transaction, DebitCard, CurrentAccountApplication, LoanApplication, LoanAccount, FixedDeposit, RecurringDeposit, Investment, CreditCard, CreditCardTransaction
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.exc import SQLAlchemyError
from flask_mail import Mail, Message
from email.message import EmailMessage
from sqlalchemy import func
from flask import send_file
from investments import InvestmentManager 
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl.styles import Font
scheduler = BackgroundScheduler()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', '083r08BHHBHBIHVDBHI938')
app.config['WTF_CSRF_ENABLED'] = False  # Temporarily disable ALL CSRF
csrf = CSRFProtect(app)
csrf._exempt_views.add('home')

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587

def send_loan_approval_email(account, loan):
    """Send loan approval confirmation email"""
    subject = f"Loan Approved - {loan.loan_id}"
    body = f"""
Dear {account.name},

Congratulations! Your loan application has been approved.

Loan Details:
-------------
Loan ID: {loan.loan_id}
Loan Type: {loan.loan_type}
Loan Amount: ₹{loan.loan_amount:,}
Interest Rate: {loan.interest_rate}% p.a.
Tenure: {loan.tenure_months} months
Monthly EMI: ₹{loan.monthly_emi:,}
First EMI Date: {loan.next_emi_date.strftime('%d-%m-%Y')}

The loan amount has been credited to your account.

Thank you for choosing YourBank!

Regards,
YourBank Loan Team
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending loan approval email: {str(e)}")

def send_loan_rejection_email(account, loan_type, reason):
    """Send loan rejection notification email"""
    subject = "Loan Application Status"
    body = f"""
Dear {account.name},

We regret to inform you that your {loan_type} loan application could not be approved at this time.

Reason: {reason}

You may reapply after 30 days or contact our customer service for more information.

Thank you for your interest in YourBank.

Regards,
YourBank Loan Team
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending loan rejection email: {str(e)}")

def send_loan_closure_email(account, loan):
    """Send loan closure confirmation email"""
    subject = f"Loan Closed - {loan.loan_id}"
    body = f"""
Dear {account.name},

Your loan has been successfully closed.

Closure Details:
----------------
Loan ID: {loan.loan_id}
Loan Type: {loan.loan_type}
Original Amount: ₹{loan.loan_amount:,}
Total Amount Paid: ₹{loan.total_paid:,}
Closure Date: {datetime.now().strftime('%d-%m-%Y')}

Thank you for your timely payments. We look forward to serving you again!

Regards,
YourBank Loan Team
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending loan closure email: {str(e)}")

def send_credit_card_application_email(account, card_type):
    """Send credit card application confirmation email"""
    subject = "Credit Card Application Received"
    body = f"""
Dear {account.name},

Thank you for applying for the {card_type} Credit Card.

Your application is being processed and you will receive a confirmation within 3-5 business days.

Application Details:
--------------------
Card Type: {card_type}
Application Date: {datetime.now().strftime('%d-%m-%Y')}
Account Number: {mask_acno(account.acno)}

We will notify you once your application is approved.

Regards,
YourBank Credit Card Team
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending credit card application email: {str(e)}")

def send_credit_card_approval_email(account, credit_card, card_password):
    """Send credit card approval and details email"""
    subject = f"🎉 Credit Card Approved - {credit_card.card_type.title()} Card"
    body = f"""Dear {account.name},

Congratulations! Your {credit_card.card_type.title()} Credit Card has been APPROVED and is ready to use!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CARD DETAILS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Card Type: {credit_card.card_type.title()} Card
Card Number: {credit_card.card_number}
Card Holder: {account.name}
Expiry Date: {credit_card.expiry_date}
CVV: {credit_card.cvv}
Credit Limit: ₹{credit_card.credit_limit:,}
Available Credit: ₹{credit_card.available_credit:,}
Card Password: {card_password}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
IMPORTANT SECURITY TIPS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✓ Never share your CVV or Card Password with anyone
✓ Change your password immediately after first use
✓ Monitor your transactions regularly
✓ Report any suspicious activity to us immediately
✓ Keep your card details safe and secure

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CARD BENEFITS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Cashback on all spends
• Reward points program
• Exclusive discounts and offers
• Travel benefits and insurance
• 24/7 customer support

Your card is now active and ready to use. You can make purchases, pay bills, and manage your credit online through our mobile app.

Thank you for choosing YourBank!

Best regards,
YourBank Credit Card Team
📞 1-800-YOURBANK
🌐 www.yourbank.com
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending credit card approval email: {str(e)}")

def send_credit_card_transaction_email(account, transaction):
    """Send credit card transaction notification"""
    subject = f"Credit Card Transaction - ₹{transaction.amount:,}"
    body = f"""
Dear {account.name},

A transaction has been made on your credit card ending in {transaction.credit_card.card_number[-4:]}.

Transaction Details:
--------------------
Amount: ₹{transaction.amount:,}
Merchant: {transaction.merchant}
Date & Time: {transaction.transaction_date.strftime('%d-%m-%Y %H:%M:%S')}
Available Credit: ₹{transaction.credit_card.available_credit:,}

If you did not make this transaction, please contact us immediately.

Regards,
YourBank Credit Card Team
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending credit card transaction email: {str(e)}")

def send_credit_card_payment_email(account, credit_card, payment_amount):
    """Send credit card payment confirmation email"""
    subject = "Credit Card Payment Received"
    body = f"""
Dear {account.name},

Your credit card payment has been received successfully.

Payment Details:
----------------
Card ending in: {credit_card.card_number[-4:]}
Payment Amount: ₹{payment_amount:,}
Available Credit: ₹{credit_card.available_credit:,}
Payment Date: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}

Thank you for your payment!

Regards,
YourBank Credit Card Team
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending credit card payment email: {str(e)}")

def send_interest_credit_email(account, interest_amount, account_type='SAVINGS'):
    """Send interest credit notification email"""
    subject = f"Interest Credited - ₹{interest_amount:,.2f}"
    body = f"""
Dear {account.name},

Interest has been credited to your {account_type} account.

Interest Details:
-----------------
Interest Amount: ₹{interest_amount:,.2f}
Credit Date: {datetime.now().strftime('%d-%m-%Y')}
Account Number: {mask_acno(account.acno)}

Thank you for banking with us!

Regards,
YourBank Team
"""
    try:
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending interest credit email: {str(e)}")
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'garvboy21@gmail.com')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', 'your_app_password')

mail = Mail(app)
db.init_app(app)


with app.app_context():
    try:
        db.create_all()
        logger.debug("Database initialization completed") 
    except Exception as e:
        logger.error(f"Error creating database tables: {str(e)}")

@app.template_filter('format_currency')
def format_currency(value):
    """Format number as currency"""
    try:
        return f"{float(value):,.2f}"
    except (ValueError, TypeError):
        return "0.00"

def generate_acno():
    while True:
        acno = ''.join(random.choices(string.digits, k=10))
        if not Account.query.filter_by(acno=acno).first():
            return acno

@app.route('/', methods=['GET'])
@csrf.exempt 
def home():
    return render_template('home.html')
@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()}

def validate_name(name):
    if not name or len(name) < 3 or len(name) > 100:
        return False
    return bool(re.match(r'^[A-Za-z\s]+$', name))

def validate_dob(dob):
    try:
        datetime.strptime(dob, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def validate_phone(phone):
    return bool(re.match(r'^\d{10}$', phone))

def validate_balance(balance):
    try:
        balance = int(balance)
        return balance >= 10000
    except ValueError:
        return False

def validate_password(password):
    return bool(re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d).{8,}$', password))


def send_email(to, subject, body, attachment=None):
    """Send email via SMTP"""
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = os.environ.get('MAIL_USERNAME', 'garvboy21@gmail.com')
        msg['To'] = to
        msg.set_content(body)

        if attachment:
            with open(attachment, 'rb') as file:
                file_data = file.read()
                file_name = os.path.basename(attachment)
                msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(os.environ.get('MAIL_USERNAME', 'garvboy21@gmail.com'), 
                        os.environ.get('MAIL_PASSWORD', 'kewx sknr pjsk phhw'))
            server.send_message(msg)
            logger.info(f"Email sent to {to}")
    except Exception as e:
        logger.error(f"Error sending email to {to}: {str(e)}")


# ==================== HELPER FUNCTIONS ====================

def mask_acno(acno):
    """Mask account number for security"""
    if not acno:
        return ''
    acno = str(acno)
    return 'X' * (len(acno) - 4) + acno[-4:] if len(acno) > 4 else acno

def send_transaction_email(account, transaction_type, amount, other_acno=None, other_name=None, direction='credit'):
    """Generic transaction notification email"""
    try:
        if not account or not account.email:
            return
        subject = f"Transaction Alert - {transaction_type}"
        
        if transaction_type in ('TRANSFER', 'PAYMENT'):
            body = f"""Dear {account.name},

A {transaction_type.lower()} has been processed on your account.

Amount: ₹{amount:,}
Direction: {direction}
Counterparty: {other_name or other_acno or 'N/A'}
Account: {mask_acno(account.acno)}
Date: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}

If you did not authorize this, contact us immediately.

Regards,
YourBank Team
"""
        elif transaction_type in ('DEPOSIT', 'WITHDRAW', 'RECEIVE'):
            body = f"""Dear {account.name},

A {transaction_type.lower()} transaction of ₹{amount:,} was processed on your account ending {mask_acno(account.acno)} on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}.

Regards,
YourBank Team
"""
        elif transaction_type == 'GST_PAYMENT':
            body = f"""Dear {account.name},

Your GST/Tax payment of ₹{amount:,} has been processed successfully.

Transaction Date: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}
Account: {mask_acno(account.acno)}

Regards,
YourBank Team
"""
        elif transaction_type == 'REDEEM':
            body = f"""Dear {account.name},

Your points redemption of ₹{amount:,.2f} has been credited to your account.

Transaction Date: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}

Regards,
YourBank Team
"""
        else:
            body = f"""Dear {account.name},

A transaction ({transaction_type}) of ₹{amount:,} was processed on your account ending {mask_acno(account.acno)} on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}.

Regards,
YourBank Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending transaction email: {e}")

def deduct_monthly_low_balance_charge(account, amount_record):
    """Deduct monthly low balance charge for CURRENT accounts"""
    try:
        if not account or account.account_type != 'CURRENT' or not amount_record:
            return
        last_charge = Transaction.query.filter_by(acno=account.acno, transaction_type='LOW_BALANCE_CHARGE').order_by(Transaction.timestamp.desc()).first()
        if last_charge and last_charge.timestamp.year == datetime.now().year and last_charge.timestamp.month == datetime.now().month:
            return
        min_balance = 50000 if 'regular' in (account.address or '').lower() else 3500000
        if amount_record.balance < min_balance:
            charge = 500
            amount_record.balance -= charge
            txn = Transaction(acno=account.acno, transaction_type='LOW_BALANCE_CHARGE', amount=charge, balance_after=amount_record.balance, timestamp=datetime.now(), description='Monthly low balance charge')
            db.session.add(txn)
            db.session.commit()
    except Exception as e:
        app.logger.error(f"Error deducting monthly low balance charge: {e}")

def deduct_quarterly_charge(account, amount_record):
    """Deduct quarterly maintenance charge for CURRENT accounts"""
    try:
        if not account or account.account_type != 'CURRENT' or not amount_record:
            return
        last_charge = Transaction.query.filter_by(acno=account.acno, transaction_type='QUARTERLY_CHARGE').order_by(Transaction.timestamp.desc()).first()
        current_quarter = (datetime.now().month - 1) // 3 + 1
        if last_charge:
            last_quarter = (last_charge.timestamp.month - 1) // 3 + 1
            if last_charge.timestamp.year == datetime.now().year and last_quarter == current_quarter:
                return
        charge = 1000 if 'regular' in (account.address or '').lower() else 300
        amount_record.balance -= charge
        txn = Transaction(acno=account.acno, transaction_type='QUARTERLY_CHARGE', amount=charge, balance_after=amount_record.balance, timestamp=datetime.now(), description='Quarterly maintenance charge')
        db.session.add(txn)
        db.session.commit()
    except Exception as e:
        app.logger.error(f"Error deducting quarterly charge: {e}")

# ==================== EMAIL FUNCTIONS ====================


def send_fd_confirmation_email(account, fd):
    """Send Fixed Deposit confirmation email"""
    try:
        if not fd or not account:
            return
        
        principal = getattr(fd, 'principal_amount', 0) or 0
        interest_rate = getattr(fd, 'interest_rate', 0) or 0
        tenure_days = getattr(fd, 'tenure_days', 0) or 0
        maturity_amount = getattr(fd, 'maturity_amount', 0) or 0
        maturity_date = getattr(fd, 'maturity_date', datetime.now())
        
        subject = f"Fixed Deposit Created - {fd.fd_id}"
        body = f"""Dear {account.name},

Your Fixed Deposit has been created successfully.

FD Details:
-----------
FD ID: {fd.fd_id}
Principal Amount: ₹{principal:,.2f}
Interest Rate: {interest_rate}% p.a.
Tenure: {tenure_days} days
Maturity Amount: ₹{maturity_amount:,.2f}
Maturity Date: {maturity_date.strftime('%d-%b-%Y')}

Your FD will automatically mature on {maturity_date.strftime('%d %B %Y')} and the maturity amount will be credited to your account.

Thank you for choosing YourBank!

Regards,
YourBank Investments Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending FD confirmation email: {e}")

def send_rd_confirmation_email(account, rd):
    """Send Recurring Deposit confirmation email"""
    try:
        if not rd or not account:
            return
        
        monthly_amount = getattr(rd, 'monthly_amount', 0) or 0
        tenure_months = getattr(rd, 'tenure_months', 0) or 0
        interest_rate = getattr(rd, 'interest_rate', 0) or 0
        maturity_amount = getattr(rd, 'maturity_amount', 0) or 0
        maturity_date = getattr(rd, 'maturity_date', datetime.now())
        next_payment = getattr(rd, 'next_payment_date', datetime.now() + timedelta(days=30))
        total_deposits = monthly_amount * tenure_months
        
        subject = f"Recurring Deposit Created - {rd.rd_id}"
        body = f"""Dear {account.name},

Your Recurring Deposit has been created successfully.

RD Details:
-----------
RD ID: {rd.rd_id}
Monthly Amount: ₹{monthly_amount:,.2f}
Tenure: {tenure_months} months
Total Amount to be Deposited: ₹{total_deposits:,.2f}
Interest Rate: {interest_rate}% p.a.
Expected Maturity Amount: ₹{maturity_amount:,.2f}
First Debit Date: {next_payment.strftime('%d-%b-%Y')}
Maturity Date: {maturity_date.strftime('%d-%b-%Y')}

Your monthly installment of ₹{monthly_amount:,.2f} will be auto-debited from your account on the 1st of every month.

Important:
- Please ensure sufficient balance in your account for monthly deductions
- In case of insufficient balance, the transaction will be retried

Thank you for choosing YourBank!

Regards,
YourBank Investments Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending RD confirmation email: {e}")

def send_investment_confirmation_email(account, investment):
    """Send investment confirmation email"""
    try:
        if not investment or not account:
            return
        
        amount = getattr(investment, 'amount', 0) or 0
        investment_type = getattr(investment, 'investment_type', 'N/A')
        purchase_price = getattr(investment, 'purchase_price', 0) or 0
        annual_return = getattr(investment, 'annual_return', 0) or 0
        invested_at = getattr(investment, 'invested_at', datetime.now())
        
        subject = f"Investment Created - {investment.investment_id}"
        body = f"""Dear {account.name},

Your investment has been created successfully.

Investment Details:
-------------------
Investment ID: {investment.investment_id}
Investment Type: {investment_type}
Amount Invested: ₹{amount:,.2f}
Purchase Price Per Unit: ₹{purchase_price:,.2f}
Expected Annual Return: {annual_return}%
Investment Date: {invested_at.strftime('%d-%b-%Y %H:%M:%S')}

Current Value Estimate: ₹{amount:,.2f}

You can track your investment performance in the Investments section of your dashboard.

Disclaimer:
Investment returns are subject to market conditions and may vary.

Thank you for choosing YourBank!

Regards,
YourBank Investments Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending investment confirmation email: {e}")

def send_investment_liquidation_email(account, investment, current_value, profit_loss):
    """Send investment liquidation confirmation email"""
    try:
        if not investment or not account:
            return
        
        original_amount = getattr(investment, 'amount', 0) or 0
        investment_type = getattr(investment, 'investment_type', 'N/A')
        return_percentage = ((profit_loss / original_amount) * 100) if original_amount > 0 else 0
        
        status = 'Profit' if profit_loss >= 0 else 'Loss'
        
        subject = f"Investment Liquidated - {investment.investment_id}"
        body = f"""Dear {account.name},

Your investment has been successfully liquidated.

Liquidation Details:
--------------------
Investment ID: {investment.investment_id}
Investment Type: {investment_type}
Original Investment Amount: ₹{original_amount:,.2f}
Liquidation Value: ₹{current_value:,.2f}
Profit/Loss: ₹{profit_loss:,.2f} ({status})
Return Percentage: {return_percentage:.2f}%
Liquidation Date: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}

Amount credited to your account: ₹{current_value:,.2f}

Thank you for choosing YourBank!

Regards,
YourBank Investments Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending investment liquidation email: {e}")

def send_fd_maturity_email(account, fd, interest):
    """Send FD maturity notification email"""
    try:
        if not fd or not account:
            return
        
        principal = getattr(fd, 'principal_amount', 0) or 0
        maturity_amount = getattr(fd, 'maturity_amount', 0) or 0
        fd_id = getattr(fd, 'fd_id', 'N/A')
        maturity_date = getattr(fd, 'maturity_date', datetime.now())
        
        subject = f"Fixed Deposit Matured - {fd_id}"
        body = f"""Dear {account.name},

Your Fixed Deposit has matured and the amount has been credited to your account.

Maturity Details:
-----------------
FD ID: {fd_id}
Principal Amount: ₹{principal:,.2f}
Interest Earned: ₹{interest:,.2f}
Total Maturity Amount: ₹{maturity_amount:,.2f}
Maturity Date: {maturity_date.strftime('%d-%b-%Y')}
Credit Date: {datetime.now().strftime('%d-%b-%Y')}

The amount has been credited to your account. You can view the transaction in your account statement.

Thank you for your continued trust in YourBank!

Regards,
YourBank Investments Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending FD maturity email: {e}")

def send_rd_maturity_email(account, rd, interest):
    """Send RD maturity notification email"""
    try:
        if not rd or not account:
            return
        
        total_paid = getattr(rd, 'total_paid', 0) or 0
        maturity_amount = getattr(rd, 'maturity_amount', 0) or 0
        rd_id = getattr(rd, 'rd_id', 'N/A')
        maturity_date = getattr(rd, 'maturity_date', datetime.now())
        tenure_months = getattr(rd, 'tenure_months', 0) or 0
        
        subject = f"Recurring Deposit Matured - {rd_id}"
        body = f"""Dear {account.name},

Your Recurring Deposit has matured and the amount has been credited to your account.

Maturity Details:
-----------------
RD ID: {rd_id}
Total Amount Deposited: ₹{total_paid:,.2f}
Tenure: {tenure_months} months
Interest Earned: ₹{interest:,.2f}
Total Maturity Amount: ₹{maturity_amount:,.2f}
Maturity Date: {maturity_date.strftime('%d-%b-%Y')}
Credit Date: {datetime.now().strftime('%d-%b-%Y')}

The amount has been credited to your account. You can view the transaction in your account statement.

Thank you for your continued trust in YourBank!

Regards,
YourBank Investments Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending RD maturity email: {e}")

def send_loan_approval_email(account, loan):
    try:
        subject = f"Loan Approved - {loan.loan_id}"
        body = f"""Dear {account.name},

Congratulations! Your loan application has been approved.

Loan ID: {loan.loan_id}
Loan Type: {loan.loan_type}
Loan Amount: ₹{loan.loan_amount:,}
Interest Rate: {loan.interest_rate}% p.a.
Tenure: {loan.tenure_months} months
Monthly EMI: ₹{loan.monthly_emi:,}

The loan amount has been credited to your account.

Regards,
YourBank Loan Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending loan approval email: {e}")

def send_loan_rejection_email(account, loan_type, reason):
    try:
        subject = "Loan Application Status"
        body = f"""Dear {account.name},

We regret to inform you that your {loan_type} loan application could not be approved.

Reason: {reason}

You may reapply after 30 days or contact our customer service.

Regards,
YourBank Loan Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending loan rejection email: {e}")

def send_loan_closure_email(account, loan):
    try:
        subject = f"Loan Closed - {loan.loan_id}"
        body = f"""Dear {account.name},

Your loan has been successfully closed.

Loan ID: {loan.loan_id}
Original Amount: ₹{loan.loan_amount:,}
Total Amount Paid: ₹{loan.total_paid:,}

Thank you for your timely payments!

Regards,
YourBank Loan Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending loan closure email: {e}")

def send_emi_payment_email(account, loan):
    try:
        subject = f"EMI Payment Successful - {loan.loan_id}"
        body = f"""Dear {account.name},

Your EMI payment of ₹{loan.monthly_emi:,} for Loan {loan.loan_id} has been processed successfully.

Outstanding Balance: ₹{loan.outstanding_balance:,}
Next EMI Date: {loan.next_emi_date.strftime('%d-%m-%Y') if loan.next_emi_date else 'N/A'}

Thank you for your payment!

Regards,
YourBank Loan Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending EMI payment email: {e}")

def send_emi_failure_email(account, loan):
    try:
        subject = f"EMI Payment Failed - {loan.loan_id}"
        body = f"""Dear {account.name},

We could not process your EMI of ₹{loan.monthly_emi:,} for Loan {loan.loan_id} due to insufficient balance.

Please top up your account to avoid penalties.

Regards,
YourBank Loan Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending EMI failure email: {e}")

def send_credit_card_application_email(account, card_type):
    try:
        subject = "Credit Card Application Received"
        body = f"""Dear {account.name},

Thank you for applying for the {card_type} Credit Card.

Your application is being processed and you will receive confirmation within 3-5 business days.

Application Date: {datetime.now().strftime('%d-%m-%Y')}

Regards,
YourBank Credit Card Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending credit card application email: {e}")


def send_credit_card_approval_email(account, credit_card, card_password):
    """Send credit card approval and details email"""
    try:
        subject = f"🎉 Credit Card Approved - {credit_card.card_type.title()} Card"
        body = f"""Dear {account.name},

Congratulations! Your {credit_card.card_type.title()} Credit Card has been APPROVED and is ready to use!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CARD DETAILS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Card Type: {credit_card.card_type.title()} Card
Card Number: {credit_card.card_number}
Card Holder: {account.name}
Expiry Date: {credit_card.expiry_date}
CVV: {credit_card.cvv}
Credit Limit: ₹{credit_card.credit_limit:,}
Available Credit: ₹{credit_card.available_credit:,}
Card Password: {card_password}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
IMPORTANT SECURITY TIPS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✓ Never share your CVV or Card Password with anyone
✓ Change your password immediately after first use
✓ Monitor your transactions regularly
✓ Report any suspicious activity to us immediately
✓ Keep your card details safe and secure

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CARD BENEFITS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Cashback on all spends
• Reward points program
• Exclusive discounts and offers
• Travel benefits and insurance
• 24/7 customer support

Your card is now active and ready to use. You can make purchases, pay bills, and manage your credit online through our mobile app.

Thank you for choosing YourBank!

Best regards,
YourBank Credit Card Team
📞 1-800-YOURBANK
🌐 www.yourbank.com
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending credit card approval email: {str(e)}")

def send_credit_card_transaction_email(account, transaction):
    try:
        subject = f"Credit Card Transaction - ₹{transaction.amount:,}"
        body = f"""Dear {account.name},

A transaction has been made on your credit card.

Amount: ₹{transaction.amount:,}
Merchant: {getattr(transaction, 'merchant', 'N/A')}
Date: {transaction.transaction_date.strftime('%d-%m-%Y %H:%M:%S')}

If you did not make this transaction, contact us immediately.

Regards,
YourBank Credit Card Team
"""
        send_email(account.email, subject, body)
    except Exception as e:
        app.logger.error(f"Error sending credit card transaction email: {e}")


def send_credit_card_payment_email(account, credit_card, payment_amount):
    """Send credit card payment confirmation email"""
    try:
        new_amount_owed = credit_card.credit_limit - credit_card.available_credit
        
        subject = f"✅ Credit Card Payment Successful - {credit_card.card_type.title()} Card"
        
        body = f"""Dear {account.name},

Your credit card payment has been successfully processed!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PAYMENT DETAILS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Transaction Date: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}
Card Type: {credit_card.card_type.title()} Card
Card Number: ****{credit_card.card_number[-4:]}
Payment Amount: ₹{payment_amount:,}
Transaction Status: COMPLETED ✓

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
UPDATED CREDIT STATUS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Credit Limit: ₹{credit_card.credit_limit:,}
Amount Owed (Used): ₹{new_amount_owed:,}
Available Credit: ₹{credit_card.available_credit:,}
Credit Utilization: {((credit_card.credit_limit - credit_card.available_credit) / credit_card.credit_limit * 100):.1f}%

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PAYMENT BREAKDOWN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Previous Amount Owed: ₹{payment_amount + new_amount_owed:,}
Payment Received: -₹{payment_amount:,}
New Amount Owed: ₹{new_amount_owed:,}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ACCOUNT IMPACT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Payment Source: Your Savings Account
Amount Deducted: ₹{payment_amount:,}
Processing Time: Immediate
Status: Successfully credited to credit card account

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
WHAT'S NEXT?
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✓ Your payment has been recorded in your transaction history
✓ Your available credit has been updated
✓ You can now make new purchases up to ₹{credit_card.available_credit:,}
✓ Your payment will reflect in your next credit card statement

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
QUICK TIPS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Keep your credit utilization below 30% for better credit health
• Make payments on time to maintain your good credit score
• Use our mobile app to track your spending and payments
• Set up payment reminders to never miss a due date

For any queries or to report unauthorized transactions, please contact us:
📞 Customer Support: 1-800-YOURBANK
📧 Email: support@yourbank.com
🌐 Website: www.yourbank.com

Thank you for banking with us!

Best regards,
YourBank Credit Card Team
"""
        
        send_email(account.email, subject, body)
        
    except Exception as e:
        app.logger.error(f"Error sending credit card payment email: {str(e)}")

# ==================== LOAN ENDPOINTS ====================

@app.route('/loans', methods=['GET'])
def loans():
    """View loan products and options"""
    if not is_session_valid():
        return redirect(url_for('login'))
    return render_template('loans.html', csrf_token=generate_csrf())

import random

@app.route('/apply_loan', methods=['GET', 'POST'])
def apply_loan():
    """Handle loan application"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            loan_type = request.form.get('loan_type', '').strip()
            loan_amount = float(request.form.get('loan_amount', 0))
            tenure_months = int(request.form.get('tenure_months', 0))
            interest_rate = float(request.form.get('interest_rate', 12.0))
            terms_agreed = request.form.get('terms_agreed')
            
            if not all([loan_type, loan_amount, tenure_months, interest_rate]):
                flash('Please fill in all required fields.', 'error')
                return render_template('apply_loan.html', csrf_token=generate_csrf())
            
            if not terms_agreed:
                flash('You must agree to the terms and conditions.', 'error')
                return render_template('apply_loan.html', csrf_token=generate_csrf())
            
            if loan_amount < 10000:
                flash('Minimum loan amount is ₹10,000.', 'error')
                return render_template('apply_loan.html', csrf_token=generate_csrf())
            
            user_acno = session.get('user_acno')
            account = Account.query.filter_by(acno=user_acno).first()
            
            if not account:
                flash('Account not found.', 'error')
                return redirect(url_for('dashboard'))
            
            monthly_rate = interest_rate / 100 / 12
            monthly_emi = (loan_amount * monthly_rate * (1 + monthly_rate) ** tenure_months) / \
                         ((1 + monthly_rate) ** tenure_months - 1)
            
            loan_id = f"LOAN{random.randint(100000, 999999)}"
            
            application = LoanApplication(
                acno=user_acno,
                loan_id=loan_id,
                loan_type=loan_type,
                loan_amount=loan_amount,
                interest_rate=interest_rate,
                tenure_months=tenure_months,
                monthly_emi=monthly_emi,
                name=account.name,
                email=account.email,
                phone=account.phone,
                dob=account.dob,
                nationality='Indian',
                status='PENDING'
            )
            
            db.session.add(application)
            db.session.commit()
            
            flash('✅ Loan application submitted successfully! Click "Process Application" to get instant approval.', 'success')
            return redirect(url_for('my_loans'))
        
        except ValueError as e:
            app.logger.error(f"ValueError in apply_loan: {str(e)}")
            flash(f'Invalid input: Please enter valid numbers.', 'error')
            return render_template('apply_loan.html', csrf_token=generate_csrf())
        except Exception as e:
            app.logger.error(f"Error submitting loan application: {str(e)}")
            flash('An error occurred while submitting your application.', 'error')
            return render_template('apply_loan.html', csrf_token=generate_csrf())
    
    return render_template('apply_loan.html', csrf_token=generate_csrf())

@app.route('/my_loans', methods=['GET'])
def my_loans():
    """View user's loans"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    acno = session.get('user_acno')
    
    loan_applications = LoanApplication.query.filter_by(acno=acno).order_by(LoanApplication.created_at.desc()).all()
    
    loan_accounts = LoanAccount.query.filter_by(acno=acno).order_by(LoanAccount.loan_id.desc()).all()
    
    applications = loan_applications
    
    return render_template('my_loans.html', 
                         loan_applications=loan_applications,
                         loan_accounts=loan_accounts,
                         applications=applications,
                         csrf_token=generate_csrf())

@app.route('/simulate_loan_processing/<int:application_id>', methods=['GET', 'POST'])
def simulate_loan_processing(application_id):
    """Simulate loan processing and approve/reject applications"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    try:
        application = LoanApplication.query.get(application_id)
        if not application or application.acno != session.get('user_acno'):
            flash('Loan application not found.', 'error')
            return redirect(url_for('my_loans'))
        
        account = Account.query.filter_by(acno=application.acno).first()
        amount_record = Amount.query.filter_by(acno=application.acno).first()
        
        if not amount_record:
            flash('Account balance record not found.', 'error')
            return redirect(url_for('my_loans'))
        
        if amount_record.balance >= 50000:
            monthly_emi = application.monthly_emi
            
            loan = LoanAccount(
                acno=application.acno,
                loan_id=application.loan_id,
                loan_type=application.loan_type,
                principal_amount=application.loan_amount,      
                disbursed_amount=application.loan_amount,     
                outstanding_balance=application.loan_amount,
                interest_rate=application.interest_rate,
                tenure_months=application.tenure_months,
                monthly_emi=int(monthly_emi),
                next_emi_date=datetime.now() + timedelta(days=30),
                status='ACTIVE'
            )
            
            amount_record.balance += application.loan_amount
            
            txn = Transaction(
                acno=application.acno,
                transaction_type='LOAN_DISBURSEMENT',
                amount=application.loan_amount,
                balance_after=amount_record.balance,
                timestamp=datetime.now(),
                description=f'Loan disbursement - {application.loan_type}'
            )
            
            application.status = 'APPROVED'
            application.approved_at = datetime.now()
            
            db.session.add(loan)
            db.session.add(txn)
            db.session.commit()
            
            send_loan_approval_email(account, loan)
            flash(f'✅ Loan approved! Amount: ₹{application.loan_amount:,.0f}, EMI: ₹{int(monthly_emi):,}', 'success')
        else:
            application.status = 'REJECTED'
            application.rejected_at = datetime.now()
            application.rejection_reason = 'Insufficient account balance (minimum ₹50,000 required)'
            db.session.commit()
            flash('❌ Loan application rejected. Minimum account balance of ₹50,000 required.', 'error')
        
        return redirect(url_for('my_loans'))
    except Exception as e:
        app.logger.error(f"Error processing loan application: {str(e)}")
        flash('An error occurred while processing your application.', 'error')
        return redirect(url_for('my_loans'))
    
@app.route('/pay_emi/<int:loan_id>', methods=['POST'])
def pay_emi(loan_id):
    """Pay EMI for a loan"""
    if not is_session_valid():
        return redirect(url_for('login'))
    
    try:
        loan = LoanAccount.query.get(loan_id)
        account = Account.query.filter_by(acno=session.get('user_acno')).first()
        
        if not loan or loan.acno != account.acno:
            flash('Loan not found or unauthorized.', 'error')
            return redirect(url_for('my_loans'))
        
        user_amount = Amount.query.filter_by(acno=account.acno).first()
        emi = loan.monthly_emi or 0
        
        if not user_amount or user_amount.balance < emi:
            send_emi_failure_email(account, loan)
            flash('Insufficient balance to pay EMI.', 'error')
            return redirect(url_for('my_loans'))
        
        user_amount.balance -= emi
        loan.outstanding_balance -= emi
        loan.total_paid = (loan.total_paid or 0) + emi
        
        if loan.next_emi_date:
            loan.next_emi_date = loan.next_emi_date + timedelta(days=30)
        
        txn = Transaction(acno=account.acno, transaction_type='EMI_PAYMENT', amount=emi, balance_after=user_amount.balance, timestamp=datetime.now())
        db.session.add(txn)
        
        if loan.outstanding_balance <= 0:
            loan.status = 'CLOSED'
            send_loan_closure_email(account, loan)
        else:
            send_emi_payment_email(account, loan)
        
        db.session.commit()
        flash(f'EMI of ₹{emi:,} paid successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error processing EMI payment: {e}")
        flash('Could not process EMI payment.', 'error')
    
    return redirect(url_for('my_loans'))


@app.route('/dashboard')
def dashboard():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    try:
        user_acno = session.get('user_acno')
        account = Account.query.filter_by(acno=user_acno).first()
        amount_record = Amount.query.filter_by(acno=user_acno).first()
        
        account_balance = amount_record.balance if amount_record else 0
        
        transactions = Transaction.query.filter_by(acno=user_acno).order_by(
            Transaction.timestamp.desc()
        ).limit(10).all()
        
        from datetime import datetime, timedelta
        thirty_days_ago = datetime.now() - timedelta(days=30)
        monthly_txns = Transaction.query.filter_by(acno=user_acno).filter(
            Transaction.timestamp >= thirty_days_ago,
            Transaction.transaction_type.in_(['WITHDRAW', 'TRANSFER', 'PAYMENT'])
        ).all()
        monthly_spending = sum(txn.amount for txn in monthly_txns)
        
        debit_cards = DebitCard.query.filter_by(account_id=account.id).all() if account else []
        
        return render_template('dashboard.html', 
                             username=session.get('user_name'),
                             account=account,
                             account_balance=account_balance,
                             transactions=transactions,
                             monthly_spending=monthly_spending,
                             debit_cards=debit_cards,
                             csrf_token=generate_csrf())
    except Exception as e:
        app.logger.error(f"Error loading dashboard: {str(e)}")
        flash('An error occurred while loading the dashboard.', 'error')
        return redirect(url_for('login'))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        dob = request.form.get('dob', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        address = request.form.get('address', '').strip()
        opening_balance = request.form.get('opening_balance', '')
        password = request.form.get('password', '')
        account_type = 'SAVINGS'

        errors = []
        if not validate_name(name):
            errors.append('Name must contain only letters and be at least 3 characters long.')
        if not validate_dob(dob):
            errors.append('Date of Birth must be in YYYY-MM-DD format.')
        if not validate_phone(phone):
            errors.append('Phone number must be exactly 10 digits.')
        if not email or not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
            errors.append('Please enter a valid email address.')
        if not address or len(address) < 5:
            errors.append('Please enter a valid address.')
        if not validate_balance(opening_balance):
            errors.append('Opening balance must be at least ₹10,000.')
        if not validate_password(password):
            errors.append('Password must be at least 8 characters with 1 uppercase, 1 lowercase, and 1 number.')

        if errors:
            for error in errors:
                flash(error, 'error')
            return render_template('signup.html', account_type=account_type)

        try:
            opening_balance = int(opening_balance)
            hashed_password = generate_password_hash(password)
            acno = generate_acno()

            def gen_debit_card_password():
                chars = string.ascii_letters + string.digits
                while True:
                    pwd = ''.join(random.choices(chars, k=10))
                    if (re.search(r'[A-Z]', pwd) and re.search(r'[a-z]', pwd) and re.search(r'\d', pwd)):
                        return pwd
            
            debit_card_password_plain = gen_debit_card_password()
            hashed_debit_card_password = generate_password_hash(debit_card_password_plain)

            new_account = Account(
                name=name,
                acno=acno,
                dob=dob,
                phone=phone,
                email=email,
                address=address,
                opening_balance=opening_balance,
                password=hashed_password,
                account_type=account_type
            )
            new_amount = Amount(acno=acno, balance=opening_balance)

            new_transaction = Transaction(
                acno=acno,
                transaction_type='DEPOSIT',
                amount=opening_balance,
                balance_after=opening_balance,
                timestamp=datetime.now()
            )

            db.session.add(new_account)
            db.session.add(new_amount)
            db.session.add(new_transaction)
            db.session.flush()  

            debit_card_number = ''.join(random.choices(string.digits, k=16))
            debit_card_expiry = f"{datetime.now().month:02d}/{datetime.now().year + 5}"
            debit_card_cvv = ''.join(random.choices(string.digits, k=3))

            new_debit_card = DebitCard(
                card_number=debit_card_number,
                account_id=new_account.id,  
                expiry_date=debit_card_expiry,
                cvv=debit_card_cvv,
                password_hash=hashed_debit_card_password
            )

            db.session.add(new_debit_card)
            db.session.commit()  

            try:
                send_email(email, 'Savings Account Created Successfully',
f'''Hi {name},

Welcome to YourBank! Your Savings Account has been created successfully.

Your Account Details:
Account Number: {acno}
Name: {name}
Account Type: {account_type}
Opening Balance: ₹{opening_balance}

Debit Card Details:
Card Number: {debit_card_number}
Expiry Date: {debit_card_expiry}
CVV: {debit_card_cvv}
Debit Card Password: {debit_card_password_plain}

Please keep this information safe.

Regards,
YourBank Team
''')
            except Exception as e:
                app.logger.error(f"Error sending email: {str(e)}")
                flash('Account created, but failed to send confirmation email.', 'warning')
            
            flash('Account created successfully! Please login.', 'success')
            return redirect(url_for('login'))
            
        except SQLAlchemyError as e:
            db.session.rollback()
            app.logger.error(f"Database error during signup: {str(e)}")
            flash('An error occurred while creating your account.', 'error')
        except Exception as e:
            app.logger.error(f"Unexpected error during signup: {str(e)}")
            flash('An unexpected error occurred.', 'error')

    return render_template('signup.html', account_type='SAVINGS')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        acno = request.form.get('acno', '').strip()
        password = request.form.get('password', '').strip()
        login_type = request.form.get('login_type', 'savings').strip()

        if not acno or not password:
            flash('Please enter both account number and password.', 'error')
            return render_template('login.html')

        try:
            if login_type == 'current':
                user = Account.query.filter_by(acno=acno, account_type='CURRENT').first()
            else:
                user = Account.query.filter_by(acno=acno, account_type='SAVINGS').first()

            if user and check_password_hash(user.password, password):
                if user.two_factor_enabled:
                    session['temp_acno'] = acno
                    flash('Please verify your 2FA code.', 'info')
                    return redirect(url_for('verify_2fa'))
                else:
                    session['user_acno'] = acno
                    session['user_name'] = user.name
                    session['account_type'] = user.account_type
                    session['last_activity'] = datetime.now().timestamp()
                    flash(f'Welcome {user.name}!', 'success')
                    
                    if user.account_type == 'CURRENT':
                        return redirect(url_for('current_account_dashboard'))
                    else:
                        return redirect(url_for('dashboard'))
            else:
                flash('Invalid account number or password.', 'error')
                
        except Exception as e:
            app.logger.error(f"Login error: {str(e)}")
            flash('An error occurred during login.', 'error')

    return render_template('login.html')


def is_session_valid():
    """Check if user session is valid"""
    if 'user_acno' not in session:
        return False
    last_activity = session.get('last_activity')
    if not last_activity:
        return False
    now = datetime.now().timestamp()
    if now - last_activity > 1800:  
        return False
    session['last_activity'] = now
    return True

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        acno = request.form.get('acno', '').strip()
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not acno or not new_password:
            flash('Please fill all fields.', 'error')
            return render_template('forgot_password.html')
            
        if new_password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('forgot_password.html')
            
        if not validate_password(new_password):
            flash('Password must be at least 8 characters with 1 uppercase, 1 lowercase, and 1 number.', 'error')
            return render_template('forgot_password.html')

        try:
            user = Account.query.filter_by(acno=acno).first()

            if user:
                user.password = generate_password_hash(new_password)
                db.session.commit()
                flash('Password updated successfully. Please login.', 'success')
                return redirect(url_for('login'))
            else:
                flash('Account not found.', 'error')
        except SQLAlchemyError as e:
            db.session.rollback()
            flash('An error occurred. Please try again.', 'error')
            app.logger.error(f"Database error during password reset: {str(e)}")

    return render_template('forgot_password.html')

@app.route('/change_debit_card_password/<int:card_id>', methods=['POST'])
def change_debit_card_password(card_id):
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    form_token = request.form.get('csrf_token')
    try:
        from flask_wtf.csrf import validate_csrf, CSRFError
        validate_csrf(form_token)
    except CSRFError:
        flash('Invalid or missing CSRF token. Please try again.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    new_password = request.form.get('new_debit_card_password', '')
    old_password = request.form.get('old_debit_card_password', '')
    if not new_password or not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d).{8,}$', new_password):
        flash('Password must be at least 8 characters with 1 uppercase, 1 lowercase, and 1 number.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    try:
        debit_card = DebitCard.query.get(card_id)
        if not debit_card:
            flash('Debit card not found.', 'error')
            return redirect(url_for('dashboard'))

        account = Account.query.filter_by(acno=session.get('user_acno')).first()
        if not account or debit_card.account_id != account.id:
            flash('Unauthorized action.', 'error')
            return redirect(url_for('dashboard'))

        if not check_password_hash(debit_card.password_hash, old_password):
            flash('Old debit card password is incorrect.', 'error')
            return redirect(request.referrer or url_for('dashboard'))

        debit_card.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash('Debit card password changed successfully.', 'success')
        return redirect(url_for('debit_card_view', card_id=card_id))
    except Exception as e:
        app.logger.error(f"Error changing debit card password: {str(e)}")
        flash('An error occurred while changing the password.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

@app.route('/deposit', methods=['GET', 'POST'])
def deposit():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            amount = request.form.get('amount', '')
            if not amount or not amount.isdigit() or int(amount) <= 0:
                flash('Please enter a valid positive amount.', 'error')
                return render_template('deposit.html')
            amount = int(amount)
            account = Account.query.filter_by(acno=session['user_acno']).first()  
            user_amount = Amount.query.filter_by(acno=session['user_acno']).first()
            if not user_amount:
                flash('Account not found.', 'error')
                if account and account.account_type == 'CURRENT':
                    return redirect(url_for('current_account_dashboard'))
                else:
                    return redirect(url_for('dashboard'))
            new_balance = user_amount.balance + amount
            user_amount.balance = new_balance
            new_transaction = Transaction(
                acno=session['user_acno'],
                transaction_type='DEPOSIT',
                amount=amount,
                balance_after=new_balance,
                timestamp=datetime.now()
            )
            db.session.add(new_transaction)
            db.session.commit()
            send_transaction_email(account, 'DEPOSIT', amount, direction='credit')
            flash(f'₹{amount} deposited successfully! New balance: ₹{new_balance}', 'success')
            if account and account.account_type == 'CURRENT':
                return redirect(url_for('current_account_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        except Exception as e:
            flash('An unexpected error occurred.', 'error')
            app.logger.error(f"Error during deposit: {str(e)}")
    return render_template('deposit.html')

@app.route('/withdraw', methods=['GET', 'POST'])
def withdraw():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            amount = request.form.get('amount', '')
            if not amount or not amount.isdigit() or int(amount) <= 0:
                flash('Please enter a valid positive amount.', 'error')
                return render_template('withdraw.html')
            amount = int(amount)
            account = Account.query.filter_by(acno=session['user_acno']).first()
            user_amount = Amount.query.filter_by(acno=session['user_acno']).first()
            if not user_amount:
                flash('Account not found.', 'error')
                if account and account.account_type == 'CURRENT':
                    return redirect(url_for('current_account_dashboard'))
                else:
                    return redirect(url_for('dashboard'))
            if user_amount.balance < amount:
                flash(f'Insufficient balance! Your current balance is ₹{user_amount.balance}.', 'error')
                return render_template('withdraw.html')
            if amount > 50000:
                flash('Withdrawal limit is ₹50,000 per transaction.', 'error')
                return render_template('withdraw.html')
            new_balance = user_amount.balance - amount
            user_amount.balance = new_balance
            new_transaction = Transaction(
                acno=session['user_acno'],
                transaction_type='WITHDRAW',
                amount=amount,
                balance_after=new_balance,
                timestamp=datetime.now()
            )
            db.session.add(new_transaction)
            db.session.commit()
            send_transaction_email(account, 'WITHDRAW', amount, direction='debit')
            flash(f'₹{amount} withdrawn successfully! Remaining balance: ₹{new_balance}', 'success')
            if account and account.account_type == 'CURRENT':
                return redirect(url_for('current_account_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        except Exception as e:
            flash('An unexpected error occurred.', 'error')
            app.logger.error(f"Error during withdrawal: {str(e)}")
    return render_template('withdraw.html')

@app.route('/balance', methods=['GET', 'POST'])
def balance():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            user_amount = Amount.query.filter_by(acno=session['user_acno']).first()
            if user_amount:
                flash(f'Your current balance is: ₹{user_amount.balance}', 'info')
            else:
                flash('Error retrieving balance information.', 'error')
        except Exception as e:
            flash('An error occurred while checking balance.', 'error')
            app.logger.error(f"Error checking balance: {str(e)}")
        account = Account.query.filter_by(acno=session['user_acno']).first()
        if account and account.account_type == 'CURRENT':
            return redirect(url_for('current_account_dashboard'))
        else:
            return redirect(url_for('dashboard'))
    return render_template('balance.html')

@app.route('/transfer', methods=['GET', 'POST'])
def transfer():
    try:
        if request.method == 'POST':
            from_acno = request.form['from_acno']
            to_acno = request.form['to_acno']
            amount = float(request.form['amount'])  

            sender = Amount.query.filter_by(acno=from_acno).first()
            receiver = Amount.query.filter_by(acno=to_acno).first()

            if not sender or not receiver:
                return render_template('transfer_money.html', message="Invalid account number(s).", csrf_token=generate_csrf())

            if from_acno == to_acno:
                return render_template('transfer_money.html', message="Cannot transfer to the same account.", csrf_token=generate_csrf())

            if sender.balance < amount:
                return render_template('transfer_money.html', message="Insufficient balance.", csrf_token=generate_csrf())

            sender.balance -= amount
            receiver.balance += amount

            sender_txn = Transaction(
                acno=from_acno,
                transaction_type='TRANSFER',
                amount=-amount,
                balance_after=sender.balance
            )

            receiver_txn = Transaction(
                acno=to_acno,
                transaction_type='RECEIVE',
                amount=amount,
                balance_after=receiver.balance
            )

            db.session.add_all([sender_txn, receiver_txn])
            db.session.commit()

            sender_account = Account.query.filter_by(acno=from_acno).first()
            receiver_account = Account.query.filter_by(acno=to_acno).first()

            send_transaction_email(sender_account, 'TRANSFER', amount, other_acno=receiver_account.acno, other_name=receiver_account.name, direction='debit')
            send_transaction_email(receiver_account, 'RECEIVE', amount, other_acno=sender_account.acno, other_name=sender_account.name, direction='credit')
            return render_template('transfer_money.html',
                                   message=f"₹{amount} transferred successfully to account {to_acno}.",
                                   csrf_token=generate_csrf())

        return render_template('transfer_money.html', csrf_token=generate_csrf(), acno=session.get('user_acno'))

    except Exception as e:
        print("Transfer error:", e)
        return f"<h3>Error during transfer:</h3><pre>{e}</pre>"

@app.route('/account_details', methods=['GET', 'POST'])
def account_details():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            user = Account.query.filter_by(acno=session['user_acno']).first()
            if user:
                return render_template('account_info.html', account=user)
            else:
                flash('Account information not found.', 'error')
                if Account.query.filter_by(acno=session['user_acno']).first().account_type == 'CURRENT':
                    return redirect(url_for('current_account_dashboard'))
                else:
                    return redirect(url_for('dashboard'))
        except Exception as e:
            flash('An error occurred while retrieving account details.', 'error')
            app.logger.error(f"Error retrieving account details: {str(e)}")
            if Account.query.filter_by(acno=session['user_acno']).first().account_type == 'CURRENT':
                return redirect(url_for('current_account_dashboard'))
            else:
                return redirect(url_for('dashboard'))
    return render_template('account_details.html')

@app.route('/transactions')
def transactions():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
        
    try:
        transactions = Transaction.query.filter_by(acno=session['user_acno']).order_by(
            Transaction.timestamp.desc()).limit(10).all()
        return render_template('transactions.html', transactions=transactions)
    except Exception as e:
        flash('An error occurred while retrieving transaction history.', 'error')
        app.logger.error(f"Error retrieving transaction history: {str(e)}")
        if Account.query.filter_by(acno=session['user_acno']).first().account_type == 'CURRENT':
            return redirect(url_for('current_account_dashboard'))
        else:
            return redirect(url_for('dashboard'))

@app.route('/close_account', methods=['GET', 'POST'])
def close_account():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            from flask_wtf.csrf import validate_csrf, CSRFError
            csrf_token = request.form.get('csrf_token', '')
            try:
                validate_csrf(csrf_token)
            except CSRFError:
                flash('Invalid or missing CSRF token. Please try again.', 'error')
                return render_template('close_account.html', csrf_token=generate_csrf())

            password = request.form.get('password', '')
            user = Account.query.filter_by(acno=session['user_acno']).first()
            if not user or not check_password_hash(user.password, password):
                flash('Invalid password. Account closure aborted.', 'error')
                return render_template('close_account.html', csrf_token=generate_csrf())

            acno = session['user_acno']
            amount_record = Amount.query.filter_by(acno=acno).first()
            if amount_record and amount_record.balance > 0:
                flash(f'Please withdraw your remaining balance of ₹{amount_record.balance} before closing the account.', 'error')
                return redirect(url_for('withdraw'))

            try:
                Transaction.query.filter_by(acno=acno).delete()
                Amount.query.filter_by(acno=acno).delete()
                Account.query.filter_by(acno=acno).delete()
                db.session.commit()

                session.clear()
                flash('Account closed successfully.', 'success')
                return redirect(url_for('home'))
            except SQLAlchemyError as e:
                db.session.rollback()
                flash('An error occurred while closing your account. Please try again.', 'error')
                app.logger.error(f"Database error during account closure: {str(e)}")
        except Exception as e:
            flash('An unexpected error occurred.', 'error')
            app.logger.error(f"Error during account closure: {str(e)}")

    from flask_wtf.csrf import generate_csrf
    return render_template('close_account.html', csrf_token=generate_csrf())

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))

@app.errorhandler(404)
def not_found_error(error):
    flash('Page not found.', 'error')
    return redirect(url_for('home'))

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    flash('An internal server error occurred.', 'error')
    logger.error(f"Internal error: {str(error)}")
    return redirect(url_for('home'))

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    logger.warning(f"CSRF Error: {str(e)}")
    if request.method == 'GET':
        return redirect(url_for('home'))
    flash('The form you submitted is expired or invalid. Please try again.', 'error')
    return redirect(request.referrer or url_for('home'))

@app.route('/create_debit_card', methods=['POST'])
def create_debit_card():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    user_acno = session.get('user_acno')
    account = Account.query.filter_by(acno=user_acno).first()
    if not account:
        flash('Account not found.', 'error')
        return redirect(url_for('dashboard'))
    card_number = ''.join(random.choices(string.digits, k=16))
    expiry_date = f"{datetime.now().month:02d}/{datetime.now().year + 5}"
    cvv = ''.join(random.choices(string.digits, k=3))
    def gen_debit_card_password():
        chars = string.ascii_letters + string.digits
        while True:
            pwd = ''.join(random.choices(chars, k=10))
            if (re.search(r'[A-Z]', pwd) and re.search(r'[a-z]', pwd) and re.search(r'\d', pwd)):
                return pwd
    debit_card_password_plain = gen_debit_card_password()
    hashed_debit_card_password = generate_password_hash(debit_card_password_plain)
    new_debit_card = DebitCard(
        card_number=card_number,
        account_id=account.id,
        expiry_date=expiry_date,
        cvv=cvv,
        password_hash=hashed_debit_card_password
    )
    try:
        db.session.add(new_debit_card)
        db.session.commit()
        flash('Debit card created successfully. Password: {}'.format(debit_card_password_plain), 'success')
    except SQLAlchemyError as e:
        db.session.rollback()
        flash('Failed to create debit card.', 'error')
        app.logger.error(f"Error creating debit card: {str(e)}")
        if Account.query.filter_by(acno=session['user_acno']).first().account_type == 'CURRENT':
            return redirect(url_for('current_account_dashboard'))
        else:
            return redirect(url_for('dashboard'))
from flask_wtf.csrf import generate_csrf, validate_csrf, CSRFError

@app.route('/debit_card/<int:card_id>', methods=['GET', 'POST'])
def debit_card_view(card_id):
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    try:
        debit_card = DebitCard.query.get(card_id)
        user_acno = session.get('user_acno')
        account = Account.query.filter_by(acno=user_acno).first()
        if not debit_card:
            app.logger.error(f"Debit card with id {card_id} not found.")
            flash('Debit card not found.', 'error')
            if Account.query.filter_by(acno=session['user_acno']).first().account_type == 'CURRENT':
                return redirect(url_for('current_account_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        if debit_card.account_id != account.id:
            app.logger.warning(f"Unauthorized access attempt to card {card_id} by account {account.id}")
            flash('Unauthorized access to debit card.', 'error')
            if Account.query.filter_by(acno=session['user_acno']).first().account_type == 'CURRENT':
                return redirect(url_for('current_account_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        transactions = Transaction.query.filter_by(acno=account.acno).order_by(Transaction.timestamp.desc()).all()
        if request.method == 'POST':
            try:
                csrf_token = request.form.get('csrf_token', '')
                validate_csrf(csrf_token)
            except CSRFError:
                flash('Invalid or missing CSRF token. Please try again.', 'error')
                return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
            transaction_type = request.form.get('transaction_type')
            amount = request.form.get('amount', '').strip()
            redeem_points = request.form.get('redeem_points', '').strip()
            user_amount = Amount.query.filter_by(acno=account.acno).first()
            if transaction_type == 'WITHDRAW':
                if not amount.isdigit() or int(amount) <= 0:
                    flash('Please enter a valid positive amount.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                amount = int(amount)
                if user_amount.balance < amount:
                    flash('Insufficient balance.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                user_amount.balance -= amount
                new_transaction = Transaction(
                    acno=account.acco,
                    transaction_type='WITHDRAW',
                    amount=amount,
                    balance_after=user_amount.balance,
                    timestamp=datetime.now()
                )
                db.session.add(new_transaction)
                db.session.commit()
                send_transaction_email(account, 'WITHDRAW', amount, direction='debit')
                flash(f'Withdrawal of ₹{amount} successful!', 'success')
                return redirect(url_for('debit_card_view', card_id=card_id))
            elif transaction_type == 'PAYMENT':
                to_acno = request.form.get('to_acno', '').strip()
                if not to_acno or not re.match(r'^\d{10}$', to_acno):
                    flash('Please enter a valid 10-digit account number to pay.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                if not amount.isdigit() or int(amount) <= 0:
                    flash('Please enter a valid positive amount.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                amount = int(amount)
                if user_amount.balance < amount:
                    flash('Insufficient balance for payment.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                if to_acno == account.acno:
                    flash('Cannot make a payment to your own account.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                receiver_amount = Amount.query.filter_by(acno=to_acno).first()
                if not receiver_amount:
                    flash('Recipient account not found.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                receiver_account = Account.query.filter_by(acno=to_acno).first()
                if not receiver_account:
                    flash('Recipient account not found.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                user_amount.balance -= amount
                receiver_amount.balance += amount
                if amount >= 10000:
                    points_earned = 170 * (amount // 10000) + ((amount % 10000) // 100)
                else:
                    points_earned = amount // 100
                account.points += points_earned
                new_transaction = Transaction(
                    acno=account.acno,
                    transaction_type='PAYMENT',
                    amount=amount,
                    balance_after=user_amount.balance,
                    timestamp=datetime.now()
                )
                receiver_transaction = Transaction(
                    acno=to_acno,
                    transaction_type='RECEIVE',
                    amount=amount,
                    balance_after=receiver_amount.balance,
                    timestamp=datetime.now()
                )
                db.session.add(new_transaction)
                db.session.add(receiver_transaction)
                db.session.commit()
                send_transaction_email(account, 'PAYMENT', amount, other_acno=to_acno, other_name=receiver_account.name, direction='debit')
                flash(f'Payment of ₹{amount} to {to_acno} successful! You earned {points_earned} points.', 'success')
                return redirect(url_for('debit_card_view', card_id=card_id))
            elif transaction_type == 'REDEEM':
                if not redeem_points.isdigit() or int(redeem_points) <= 0:
                    flash('Please enter a valid number of points to redeem.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                redeem_points = int(redeem_points)
                if account.points < redeem_points:
                    flash('Not enough points to redeem.', 'error')
                    return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
                cash = redeem_points * 0.25
                user_amount.balance += cash
                account.points -= redeem_points
                db.session.commit()
                send_transaction_email(account, 'REDEEM', cash, direction='credit')
                flash(f'Redeemed {redeem_points} points for ₹{cash:.2f}.', 'success')
                return redirect(url_for('debit_card_view', card_id=card_id))
            else:
                flash('Invalid transaction type.', 'error')
                return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
        return render_template('debit_card.html', debit_card=debit_card, transactions=transactions, csrf_token=generate_csrf(), points=account.points)
    except Exception as e:
        app.logger.error(f"Error in debit card view: {str(e)}")
        flash('An error occurred. Please try again.', 'error')
        if Account.query.filter_by(acno=session['user_acno']).first().account_type == 'CURRENT':
            return redirect(url_for('current_account_dashboard'))
        else:
            return redirect(url_for('dashboard'))
        
@app.route('/debit_card/<int:card_id>/change_password', methods=['GET'])
def debit_card_change_password(card_id):
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    debit_card = DebitCard.query.get(card_id)
    user_acno = session.get('user_acno')
    account = Account.query.filter_by(acno=user_acno).first()
    if not debit_card or not account or debit_card.account_id != account.id:
        flash('Unauthorized access to debit card.', 'error')
        return redirect(url_for('dashboard'))
    from flask_wtf.csrf import generate_csrf
    return render_template('change_debit_card_password.html', debit_card=debit_card, csrf_token=generate_csrf())

@app.route('/apply_current_account', methods=['GET'])
@csrf.exempt
def apply_current_account():
    """Redirect to current account application selection"""
    return render_template('apply_current_account.html', csrf_token=generate_csrf())

@app.route('/current_account_info/<account_type>')
def current_account_info(account_type):
    if account_type not in ['regular', 'premium']:
        abort(404)
    return render_template('current_account_info.html', account_type=account_type)

@app.route('/current_account_apply/<account_type>', methods=['GET', 'POST'])
def current_account_apply(account_type):
    if account_type not in ['regular', 'premium']:
        abort(404)

    business_types = [
        'Business',
        'Sole Proprietorship',
        'Private Limited Company',
        'Public Limited Company',
        'Trust',
        'Association',
        'Partnership',
        'LLP'
    ]

    business_type_min_turnover = {
        'Business': 1000000,
        'Sole Proprietorship': 500000,
        'Private Limited Company': 2000000,
        'Public Limited Company': 5000000,
        'Trust': 100000,
        'Association': 100000,
        'Partnership': 500000,
        'LLP': 1000000,
    }

    if request.method == 'POST':
        try:
            csrf_token = request.form.get('csrf_token', '')
            validate_csrf(csrf_token)

            name = request.form.get('name', '').strip()
            phone = request.form.get('phone', '').strip()
            email = request.form.get('email', '').strip()
            dob = request.form.get('dob', '').strip()
            business_type = request.form.get('business_type', '').strip()
            company_name = request.form.get('company_name', '').strip()
            turnover = request.form.get('turnover', '0').strip() 
            start_date = request.form.get('start_date', '').strip()  
            
            errors = []  
            
            if not name or len(name) < 3:
                errors.append('Please enter a valid name.')
            if not phone or not re.match(r'^\d{10}$', phone):
                errors.append('Please enter a valid 10-digit phone number.')
            if not email or not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
                errors.append('Please enter a valid email address.')
            if not dob:
                errors.append('Please enter your date of birth.')
            if not business_type:
                errors.append('Please select a business type.')
            if not company_name or len(company_name) < 3:
                errors.append('Please enter a valid company/organization name.')
            
            try:
                turnover_val = int(turnover)
            except Exception:
                errors.append('Please enter a valid turnover amount (numbers only).')
                turnover_val = 0
            
            try:
                datetime.strptime(start_date, '%Y-%m-%d')
            except Exception:
                errors.append('Please enter a valid company/organization start date (YYYY-MM-DD).')

            min_turnover = business_type_min_turnover.get(business_type, 0)
            if turnover_val < min_turnover:
                errors.append(f'Minimum annual turnover required for {business_type} is ₹{min_turnover:,}.')

            if account_type == 'premium' and turnover_val < 40000000:
                errors.append('Minimum turnover required for Premium Current Account is ₹4,00,00,000.')

            if errors:
                for error in errors:
                    flash(error, 'error')
                return render_template(
                    'current_account_apply.html',
                    account_type=account_type,
                    business_types=business_types,
                    csrf_token=generate_csrf(),
                    form=request.form
                )

            acno = generate_acno()

            application = CurrentAccountApplication(
                account_type=account_type,
                name=name,
                phone=phone,
                email=email,
                dob=dob,
                business_type=business_type,
                company_name=company_name,
                turnover=turnover_val,
                start_date=start_date,
                acno=acno
            )
            db.session.add(application)

            def gen_password():
                chars = string.ascii_letters + string.digits
                while True:
                    pwd = ''.join(random.choices(chars, k=10))
                    if (re.search(r'[A-Z]', pwd) and re.search(r'[a-z]', pwd) and re.search(r'\d', pwd)):
                        return pwd  
            auto_password = gen_password()
            opening_balance = 50000 if account_type == 'regular' else 3500000
            account = Account(
                name=name,
                acno=acno,
                dob=dob,
                phone=phone,
                email=email,
                address=company_name,
                opening_balance=opening_balance,
                password=generate_password_hash(auto_password),  
                account_type='CURRENT'
            )
            db.session.add(account)
            db.session.add(Amount(acno=acno, balance=opening_balance))

            db.session.commit()

            if account_type == 'regular':
                features = (
                    "• Monthly account balance: ₹50,000\n"
                    "• Free cash deposit up to ₹18,00,000 per month\n"
                    "• Withdraw & deposit cash at all branches\n"
                    "• Free access to secure internet banking\n"
                    "• Free 20 Demand Drafts per month\n"
                    "• Quarterly charges: ₹1,000"
                )
            else:
                features = (
                    "• Monthly account balance: ₹35,00,000\n"
                    "• Free cash deposit up to ₹5,00,00,000 per month\n"
                    "• Unlimited free cash withdrawal from home branch\n"
                    "• Unlimited free Demand Drafts/Banker Cheques\n"
                    "• Quarterly charges: ₹300"
                )

            subject = "Welcome to YourBank - Your Current Account Details"
            body = f"""Dear {name},

Congratulations and welcome to YourBank! Your {account_type.title()} Current Account has been created successfully.

Your Account Details:
----------------------
Account Number: {acno}
Account Type: {account_type.title()} Current Account
Login Password: {auto_password}

Exclusive Features for You:
{features}

Please keep this information safe. You can log in using your account number and the above password. For your security, we recommend changing your password after your first login.

Thank you for choosing YourBank. We look forward to serving your business needs!

Best regards,
YourBank Team
"""

            send_email(email, subject, body)

            flash('Your current account has been created! Account number and password have been sent to your email.', 'success')
            return redirect(url_for('login'))

        except CSRFError:
            flash('Invalid or missing CSRF token. Please try again.', 'error')
        except Exception as e:
            flash('An error occurred while submitting your application.', 'error')
            app.logger.error(f"Error in current account application: {str(e)}")

        return render_template(
            'current_account_apply.html',
            account_type=account_type,
            business_types=business_types,
            csrf_token=generate_csrf(),
            form=request.form
        )

    return render_template(
        'current_account_apply.html',
        account_type=account_type,
        business_types=business_types,
        csrf_token=generate_csrf(),
        form={}
    )

@app.route('/current_account_dashboard', methods=['GET', 'POST'])
def current_account_dashboard():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    try:
        user_acno = session.get('user_acno')
        account = Account.query.filter_by(acno=user_acno).first()
        if not account or account.account_type != 'CURRENT':
            flash('Access restricted to Current Account holders.', 'error')
            return redirect(url_for('login'))

        amount_record = Amount.query.filter_by(acno=user_acno).first()
        deduct_monthly_low_balance_charge(account, amount_record)
        deduct_quarterly_charge(account, amount_record)

        transactions = Transaction.query.filter_by(acno=user_acno).order_by(Transaction.timestamp.desc()).limit(100).all()

        from collections import defaultdict

        turnover_data = defaultdict(float)
        expense_data = defaultdict(float)
        income_data = defaultdict(float)
        for txn in transactions:
            month = txn.timestamp.strftime('%b %Y')
            if txn.transaction_type in ['DEPOSIT', 'RECEIVE']:
                turnover_data[month] += txn.amount
                income_data[month] += txn.amount
            elif txn.transaction_type in ['WITHDRAW', 'TRANSFER', 'PAYMENT']:
                turnover_data[month] -= txn.amount
                expense_data[month] += txn.amount

        alerts = []
        amount = Amount.query.filter_by(acno=user_acno).first()
        if amount and amount.balance < 10000:
            alerts.append("Low balance alert: Your balance is below ₹10,000.")
        for txn in transactions:
            if abs(txn.amount) > 100000:
                alerts.append(f"Large transaction alert: ₹{txn.amount} on {txn.timestamp.strftime('%d-%b-%Y')}.")

        offers = [
            "Business Loan: Up to ₹50,00,000 at 9.5% p.a.",
            "Overdraft Facility: Up to ₹10,00,000 with zero processing fee.",
            "Credit Card Offer: Free for first year, 2% cashback on business spends.",
            "Partner Offer: 20% off on Zoho Books subscription.",
            "Partner Offer: 15% discount on Delhivery logistics."
        ]

        important_dates = [
            {"date": "07-20-2025", "event": "GST Filing Due"},
            {"date": "07-31-2025", "event": "Quarterly TDS Payment"},
            {"date": "08-15-2025", "event": "Advance Tax Installment"},
        ]

        return render_template(
            'current_account_dashboard.html',
            account=account,
            transactions=transactions[:10],
            turnover_data=turnover_data,
            expense_data=expense_data,
            income_data=income_data,
            alerts=alerts,
            offers=offers,
            important_dates=important_dates
        )
    except Exception as e:
        app.logger.error(f"Error loading current account dashboard: {str(e)}")
        flash('An error occurred while loading the dashboard.', 'error')
        return redirect(url_for('login'))

@app.route('/bulk_payments')
def bulk_payments():
    return render_template('bulk_payments.html')

@app.route('/gst_tax_payments', methods=['GET', 'POST'])
def gst_tax_payments():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    if request.method == 'POST':
        gst_amount = request.form.get('gst_amount', '')
        if not gst_amount or not gst_amount.isdigit() or int(gst_amount) <= 0:
            flash('Please enter a valid GST amount.', 'error')
            return render_template('gst_tax_payments.html')
        gst_amount = int(gst_amount)
        account = Account.query.filter_by(acno=session['user_acno']).first()
        amount_record = Amount.query.filter_by(acno=session['user_acno']).first()
        if amount_record.balance < gst_amount:
            flash('Insufficient balance for GST payment.', 'error')
            return render_template('gst_tax_payments.html')
        amount_record.balance -= gst_amount
        txn = Transaction(
            acno=account.acno,
            transaction_type='GST_PAYMENT',
            amount=gst_amount,
            balance_after=amount_record.balance,
            timestamp=datetime.now(),
            description='GST/Tax Payment'
        )
        db.session.add(txn)
        db.session.commit()
        send_transaction_email(account, 'GST_PAYMENT', gst_amount, direction='debit')
        flash('GST/Tax payment processed and recorded successfully!', 'success')
        if account and account.account_type == 'CURRENT':
            return redirect(url_for('current_account_dashboard'))
        else:
            return redirect(url_for('dashboard'))
    return render_template('gst_tax_payments.html')

@app.route('/download_statements', methods=['POST'])
def download_statements():
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))

    format = request.form.get('format')
    user_acno = session.get('user_acno')
    account = Account.query.filter_by(acno=user_acno).first()
    transactions = Transaction.query.filter_by(acno=user_acno).order_by(Transaction.timestamp.desc()).all()

    if not account:
        flash('Account not found.', 'error')
        return redirect(url_for('dashboard'))

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['YourBank Statement'])
    writer.writerow(['Account Number:', account.acno])
    writer.writerow(['Name:', account.name])
    writer.writerow(['Account Type:', account.account_type])
    writer.writerow([])
    writer.writerow(['Date', 'Type', 'Amount', 'Balance After', 'Description'])
    for txn in transactions:
        writer.writerow([
            txn.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            txn.transaction_type,
            txn.amount,
            txn.balance_after,
            getattr(txn, 'description', '')
        ])
    output = si.getvalue()

    filename = f"statement_{user_acno}.csv"
    with open(filename, "w") as f:
        f.write(output)

    try:
        send_email(
            account.email,
            "YourBank Account Statement",
            f"Dear {account.name},\n\nPlease find attached your latest account statement.\n\nRegards,\nYourBank Team",
            attachment=filename
        )
        flash('Your statement will be sent to your registered email shortly.', 'success')
    except Exception as e:
        app.logger.error(f"Error sending statement email: {str(e)}")
        flash('Failed to send statement email.', 'error')

    try:
        os.remove(filename)
    except Exception:
        pass

    if account and account.account_type == 'CURRENT':
        return redirect(url_for('current_account_dashboard'))
    else:
        return redirect(url_for('dashboard'))

@app.route('/download_template/<template_type>')
def download_template(template_type):
    """Download sample template for bulk payments"""
    
    sample_data = [
        ['Account Number', 'Amount', 'Description', 'Reference'],
        ['1234567890', '1000.00', 'Salary Payment - John Doe', 'REF001'],
        ['2345678901', '1500.50', 'Invoice Payment - ABC Corp', 'REF002'],
        ['3456789012', '2000.00', 'Bonus Payment - Jane Smith', 'REF003']
    ]
    
    if template_type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bulk Payment Template"
        
        for row_num, row_data in enumerate(sample_data, 1):
            for col_num, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_data)
                if row_num == 1:  # Header row
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name='bulk_payment_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    elif template_type == 'csv':
        temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='')
        writer = csv.writer(temp_file)
        writer.writerows(sample_data)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name='bulk_payment_template.csv',
            mimetype='text/csv'
        )
    else:
        flash('Invalid template type requested.', 'error')
        return redirect(url_for('bulk_payments'))

@app.route('/cleanup_temp_file')
def cleanup_temp_file():
    """Clean up temporary files after download"""
    return '', 204

@app.route('/fixed_deposit', methods=['GET', 'POST'])
def fixed_deposit():
    """Create Fixed Deposit"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            amount = int(request.form.get('amount', 0))
            tenure_days = int(request.form.get('tenure', 0))
            
            if amount < 10000:
                flash('Minimum FD amount is ₹10,000.', 'error')
                return render_template('fixed_deposit.html', csrf_token=generate_csrf())
            
            account = Account.query.filter_by(acno=session['user_acno']).first()
            amount_record = Amount.query.filter_by(acno=session['user_acno']).first()
            
            if amount_record.balance < amount:
                flash('Insufficient balance.', 'error')
                return render_template('fixed_deposit.html', csrf_token=generate_csrf())
            
            amount_record.balance -= amount
            
            interest_rate = 6.5 
            maturity_amount = amount * (1 + (interest_rate/100) * (tenure_days/365))
            maturity_date = datetime.now() + timedelta(days=tenure_days)
            
            fd = FixedDeposit(
                acno=account.acno,
                fd_id=f"FD{random.randint(100000, 999999)}",
                principal_amount=amount,
                interest_rate=interest_rate,
                tenure_days=tenure_days,
                maturity_amount=maturity_amount,
                maturity_date=maturity_date,
                status='ACTIVE'
            )
            
            txn = Transaction(
                acno=account.acno,
                transaction_type='FD_CREATION',
                amount=amount,
                balance_after=amount_record.balance,
                timestamp=datetime.now()
            )
            
            db.session.add(fd)
            db.session.add(txn)
            db.session.commit()
            
            send_fd_confirmation_email(account, fd)
            flash(f'Fixed Deposit created successfully! FD ID: {fd.fd_id}', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            app.logger.error(f"Error creating FD: {str(e)}")
            flash('An error occurred while creating Fixed Deposit.', 'error')
    
    return render_template('fixed_deposit.html', csrf_token=generate_csrf())

@app.route('/recurring_deposit', methods=['GET', 'POST'])
def recurring_deposit():
    """Create Recurring Deposit"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            monthly_amount = int(request.form.get('monthly_amount', 0))
            tenure_months = int(request.form.get('tenure', 0))
            
            if monthly_amount < 1000:
                flash('Minimum monthly amount is ₹1,000.', 'error')
                return render_template('recurring_deposit.html', csrf_token=generate_csrf())
            
            account = Account.query.filter_by(acno=session['user_acno']).first()
            amount_record = Amount.query.filter_by(acno=session['user_acno']).first()
            
            if amount_record.balance < monthly_amount:
                flash('Insufficient balance for first installment.', 'error')
                return render_template('recurring_deposit.html', csrf_token=generate_csrf())
            
            amount_record.balance -= monthly_amount
            
            interest_rate = 6.0 
            total_deposits = monthly_amount * tenure_months
            maturity_amount = total_deposits * (1 + (interest_rate/100) * (tenure_months/12))
            maturity_date = datetime.now() + timedelta(days=tenure_months*30)
            
            rd = RecurringDeposit(
                acno=account.acno,
                rd_id=f"RD{random.randint(100000, 999999)}",
                monthly_amount=monthly_amount,
                interest_rate=interest_rate,
                tenure_months=tenure_months,
                maturity_amount=maturity_amount,
                maturity_date=maturity_date,
                next_payment_date=datetime.now() + timedelta(days=30),
                total_paid=monthly_amount,
                status='ACTIVE'
            )
            
            txn = Transaction(
                acno=account.acno,
                transaction_type='RD_PAYMENT',
                amount=monthly_amount,
                balance_after=amount_record.balance,
                timestamp=datetime.now()
            )
            
            db.session.add(rd)
            db.session.add(txn)
            db.session.commit()
            
            send_rd_confirmation_email(account, rd)
            flash(f'Recurring Deposit created successfully! RD ID: {rd.rd_id}', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            app.logger.error(f"Error creating RD: {str(e)}")
            flash('An error occurred while creating Recurring Deposit.', 'error')
    
    return render_template('recurring_deposit.html', csrf_token=generate_csrf())


@app.route('/investments', methods=['GET'])
def investments():
    """View all investments"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    try:
        acno = session.get('user_acno')
        if not acno:
            flash('Your session has expired. Please login again.', 'error')
            return redirect(url_for('login'))
        
        account = Account.query.filter_by(acno=acno).first()
        if not account:
            flash('Account not found.', 'error')
            return redirect(url_for('login'))
        
        investments_list = Investment.query.filter_by(acno=acno, status='ACTIVE').all()
        fds = FixedDeposit.query.filter_by(acno=acno, status='ACTIVE').all()
        rds = RecurringDeposit.query.filter_by(acno=acno, status='ACTIVE').all()
        
        total_fd = 0
        for fd in fds:
            total_fd += getattr(fd, 'principal_amount', 0) or 0
        
        total_rd = 0
        for rd in rds:
            monthly = getattr(rd, 'monthly_amount', 0) or 0
            tenure = getattr(rd, 'tenure_months', 0) or 0
            total_rd += (monthly * tenure)
        
        total_investments = 0
        for inv in investments_list:
            total_investments += getattr(inv, 'amount', 0) or 0
        
        total_returns = 0
        
        try:
            investment_manager = InvestmentManager()
            
            for inv in investments_list:
                try:
                    inv.current_value = investment_manager.get_current_value(inv)
                    inv.profit_loss = inv.current_value - inv.amount
                    total_returns += inv.profit_loss
                except Exception as e:
                    app.logger.error(f"Error calculating investment value: {e}")
                    inv.current_value = inv.amount
                    inv.profit_loss = 0
        except Exception as e:
            app.logger.error(f"Error initializing InvestmentManager: {e}")
        
        for fd in fds:
            fd_principal = getattr(fd, 'principal_amount', 0) or 0
            fd_maturity = getattr(fd, 'maturity_amount', 0) or 0
            fd_return = fd_maturity - fd_principal
            total_returns += fd_return
        
        for rd in rds:
            rd_paid = getattr(rd, 'monthly_amount', 0) * getattr(rd, 'tenure_months', 0) or 0
            rd_maturity = getattr(rd, 'maturity_amount', 0) or 0
            rd_return = rd_maturity - rd_paid
            total_returns += rd_return
        
        return render_template('investments.html', 
                             investments=investments_list,
                             fixed_deposits=fds,
                             recurring_deposits=rds,
                             total_fd=total_fd,
                             total_rd=total_rd,
                             total_investments=total_investments,
                             total_returns=total_returns,
                             csrf_token=generate_csrf())
    except Exception as e:
        app.logger.error(f"Error loading investments: {str(e)}")
        flash('An error occurred while loading investments.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/create_investment', methods=['POST'])
def create_investment():
    """Create new investment"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    try:
        investment_type = request.form.get('investment_type')
        amount = int(request.form.get('amount', 0))
        
        if amount < 5000:
            flash('Minimum investment amount is ₹5,000.', 'error')
            return redirect(url_for('investments'))
        
        account = Account.query.filter_by(acno=session['user_acno']).first()
        amount_record = Amount.query.filter_by(acno=session['user_acno']).first()
        
        if amount_record.balance < amount:
            flash('Insufficient balance.', 'error')
            return redirect(url_for('investments'))
        
        amount_record.balance -= amount
        
        investment_manager = InvestmentManager()
        purchase_price = investment_manager.get_current_price(investment_type)
        annual_return = investment_manager.get_expected_return(investment_type)
        
        quantity = amount / purchase_price if purchase_price > 0 else amount / 100
        
        investment = Investment(
            acno=account.acno,
            investment_id=f"INV{random.randint(100000, 999999)}",
            investment_type=investment_type,
            amount=amount,
            quantity=quantity,
            purchase_price=purchase_price,
            current_price=purchase_price,
            annual_return=annual_return,
            invested_at=datetime.now(),
            status='ACTIVE'
        )
        
        txn = Transaction(
            acno=account.acno,
            transaction_type='INVESTMENT',
            amount=amount,
            balance_after=amount_record.balance,
            timestamp=datetime.now(),
            description=f'Investment in {investment_type} - {quantity:.2f} units @ ₹{purchase_price:.2f}'
        )
        
        db.session.add(investment)
        db.session.add(txn)
        db.session.commit()
        
        send_investment_confirmation_email(account, investment)
        flash(f'Investment created successfully! Investment ID: {investment.investment_id}', 'success')
        
    except Exception as e:
        app.logger.error(f"Error creating investment: {str(e)}")
        flash('An error occurred while creating investment.', 'error')
    
    return redirect(url_for('investments'))

@app.route('/liquidate_investment/<int:investment_id>', methods=['POST'])
def liquidate_investment(investment_id):
    """Liquidate an investment"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    try:
        investment = Investment.query.get(investment_id)
        account = Account.query.filter_by(acno=session['user_acno']).first()
        
        if not investment or investment.acno != account.acno:
            flash('Investment not found.', 'error')
            return redirect(url_for('investments'))
        
        if investment.status != 'ACTIVE':
            flash('Investment is already liquidated.', 'error')
            return redirect(url_for('investments'))
        
        investment_manager = InvestmentManager()
        current_value = investment_manager.get_current_value(investment)
        profit_loss = current_value - investment.amount
        
        current_price = investment_manager.get_current_price(investment.investment_type)
        quantity = getattr(investment, 'quantity', 0) or 0
        
        investment.current_price = current_price
        investment.status = 'LIQUIDATED'
        investment.liquidated_at = datetime.now()
        
        amount_record = Amount.query.filter_by(acno=account.acno).first()
        amount_record.balance += current_value
        
        txn = Transaction(
            acno=account.acno,
            transaction_type='INVESTMENT_LIQUIDATION',
            amount=current_value,
            balance_after=amount_record.balance,
            timestamp=datetime.now(),
            description=f'Liquidated {quantity:.2f} units @ ₹{current_price:.2f}'
        )
        
        db.session.add(txn)
        db.session.commit()
        
        send_investment_liquidation_email(account, investment, current_value, profit_loss)
        flash(f'Investment liquidated successfully! Amount credited: ₹{current_value:,.2f}', 'success')
        
    except Exception as e:
        app.logger.error(f"Error liquidating investment: {str(e)}")
        flash('An error occurred while liquidating investment.', 'error')
    
    return redirect(url_for('investments'))
# ==================== CREDIT CARD ENDPOINTS ====================

@app.route('/apply_credit_card', methods=['GET', 'POST'])
def apply_credit_card():
    """Apply for a credit card with instant eligibility checks and approval/rejection"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            card_type = request.form.get('card_type', 'standard')
            declare_eligible = request.form.get('declare_eligible')
            terms_agreed = request.form.get('terms_agreed')
            
            if not declare_eligible or not terms_agreed:
                flash('Please declare eligibility and agree to terms.', 'error')
                return render_template('apply_credit_card.html', csrf_token=generate_csrf())
            
            account = Account.query.filter_by(acno=session['user_acno']).first()
            amount_record = Amount.query.filter_by(acno=account.acno).first()
            
            if not account or not amount_record:
                flash('Account not found.', 'error')
                return redirect(url_for('login'))
            
            existing_card = CreditCard.query.filter_by(acno=account.acno, card_type=card_type, status='ACTIVE').first()
            if existing_card:
                flash('You already have an active card of this type.', 'error')
                return redirect(url_for('apply_credit_card'))
            
            card_requirements = {
                'standard': {
                    'name': 'Standard Card',
                    'min_balance': 50000,
                    'min_annual_income': 300000,
                    'min_cibil': 700,
                    'min_age': 21,
                    'max_age': 60,
                    'credit_limit': 100000,
                    'annual_fee': 500
                },
                'gold': {
                    'name': 'Gold Card',
                    'min_balance': 200000,
                    'min_annual_income': 1000000,
                    'min_cibil': 750,
                    'min_age': 23,
                    'max_age': 65,
                    'credit_limit': 500000,
                    'annual_fee': 2500
                },
                'platinum': {
                    'name': 'Platinum Card',
                    'min_balance': 1000000,
                    'min_annual_income': 2500000,
                    'min_cibil': 800,
                    'min_age': 25,
                    'max_age': 70,
                    'credit_limit': 1000000,
                    'annual_fee': 10000
                }
            }
            
            requirements = card_requirements.get(card_type, card_requirements['standard'])
            card_name = requirements['name']
            
            eligibility_errors = {}
            
            eligibility_errors['balance'] = amount_record.balance >= requirements['min_balance']
            
            estimated_annual_income = amount_record.balance * 12 * 0.3
            eligibility_errors['income'] = estimated_annual_income >= requirements['min_annual_income']
            
            from datetime import date
            try:
                dob = datetime.strptime(account.dob, '%Y-%m-%d').date()
                today = date.today()
                age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
                
                eligibility_errors['age'] = age >= requirements['min_age'] and age <= requirements['max_age']
                
                if not eligibility_errors['age']:
                    if age < requirements['min_age']:
                        eligibility_errors['age'] = f"❌ Minimum age: {requirements['min_age']} years (Your age: {age} years)"
                    elif age > requirements['max_age']:
                        eligibility_errors['age'] = f"❌ Maximum age: {requirements['max_age']} years (Your age: {age} years)"
                else:
                    eligibility_errors['age'] = f"✅ Age: {age} years (Required: {requirements['min_age']}-{requirements['max_age']})"
            except Exception as e:
                app.logger.error(f"Error calculating age: {e}")
                eligibility_errors['age'] = "❌ Unable to verify age"
            
            all_passed = all(eligibility_errors.values())
            
            if all_passed:
                try:
                    card_number = ''.join(random.choices(string.digits, k=16))
                    expiry_month = datetime.now().month
                    expiry_year = datetime.now().year + 5
                    cvv = ''.join(random.choices(string.digits, k=3))
                    
                    credit_limit = requirements['credit_limit']
                    
                    def gen_card_password():
                        chars = string.ascii_letters + string.digits
                        while True:
                            pwd = ''.join(random.choices(chars, k=10))
                            if (re.search(r'[A-Z]', pwd) and re.search(r'[a-z]', pwd) and re.search(r'\d', pwd)):
                                return pwd
                    
                    card_password = gen_card_password()
                    
                    credit_card = CreditCard(
                        acno=account.acno,
                        card_number=card_number,
                        card_type=card_type,
                        expiry_date=f"{expiry_month:02d}/{expiry_year}",
                        cvv=cvv,
                        credit_limit=credit_limit,
                        available_credit=credit_limit,
                        current_balance=0,
                        status='ACTIVE',
                        created_at=datetime.now()
                    )
                    
                    db.session.add(credit_card)
                    db.session.commit()
                    
                    send_credit_card_approval_email(account, credit_card, card_password)
                    
                    flash_message = f"""
                    ✅ <strong>CONGRATULATIONS! Your {card_name} Credit Card has been APPROVED!</strong><br><br>
                    <strong>Card Details:</strong><br>
                    Card Number: {card_number[:4]}****{card_number[-4:]}<br>
                    Credit Limit: ₹{credit_limit:,}<br>
                    Validity: {expiry_month:02d}/{expiry_year}<br><br>
                    Complete details have been sent to your registered email.
                    """
                    flash(flash_message, 'success')
                    
                    return redirect(url_for('my_credit_cards'))
                    
                except Exception as e:
                    app.logger.error(f"Error creating credit card: {str(e)}")
                    flash('❌ An error occurred while creating your credit card.', 'error')
                    return render_template('apply_credit_card.html', csrf_token=generate_csrf())
            
            else:
                rejection_reasons = []
                if not eligibility_errors.get('balance', False):
                    rejection_reasons.append(f"Minimum account balance requirement not met (Required: ₹{requirements['min_balance']:,})")
                if not eligibility_errors.get('income', False):
                    rejection_reasons.append(f"Minimum annual income requirement not met (Required: ₹{requirements['min_annual_income']:,})")
                if not eligibility_errors.get('age', False):
                    rejection_reasons.append(f"Age requirement not met (Required: {requirements['min_age']}-{requirements['max_age']} years)")
                
                reason = "; ".join(rejection_reasons)
                
                rejection_email_body = f"""Dear {account.name},

We regret to inform you that your application for the {card_name} Credit Card could not be approved at this time.

Eligibility Assessment:
{chr(10).join([f"• {k}: {'Met' if v else 'Not met'}" for k, v in eligibility_errors.items()])}

Reasons for rejection:
{chr(10).join([f"• {r}" for r in rejection_reasons])}

<strong>What you can do:</strong>
• Improve your account balance to ₹{requirements['min_balance']:,}
• Maintain consistent income levels
• Reapply after 30 days once you meet the requirements

For any queries, please contact our customer support team.

Best regards,
YourBank Credit Card Team
"""
                
                try:
                    send_email(
                        account.email,
                        f'Credit Card Application Status - {card_name}',
                        rejection_email_body
                    )
                except Exception as e:
                    app.logger.error(f"Error sending rejection email: {e}")
                
                flash_message = f"""
                ❌ <strong>Your {card_name} Credit Card application has been REJECTED</strong><br><br>
                <strong>Eligibility Assessment:</strong><br>
                """
                for check in eligibility_errors:
                    flash_message += f"{check}<br>"
                
                flash_message += f"<br><strong>Next Steps:</strong><br>"
                for reason in rejection_reasons:
                    flash_message += f"• {reason}<br>"
                
                flash_message += f"""
                <br>A detailed rejection letter has been sent to your registered email.
                You can reapply after improving your financial profile.
                """
                
                flash(flash_message, 'error')
                return render_template('apply_credit_card.html', csrf_token=generate_csrf())
        
        except Exception as e:
            app.logger.error(f"Error processing credit card application: {str(e)}")
            flash(f'An error occurred while processing your application: {str(e)}', 'error')
    
    return render_template('apply_credit_card.html', csrf_token=generate_csrf())

@app.route('/my_credit_cards', methods=['GET'])
def my_credit_cards():
    """View user's credit cards"""
    if not is_session_valid():
        return redirect(url_for('login'))
    
    try:
        account = Account.query.filter_by(acno=session['user_acno']).first()
        credit_cards = CreditCard.query.filter_by(acno=account.acno).all()
        return render_template('my_credit_cards.html', credit_cards=credit_cards, csrf_token=generate_csrf())
    except Exception as e:
        app.logger.error(f"Error loading credit cards: {str(e)}")
        flash('An error occurred while loading credit cards.', 'error')
        return redirect(url_for('dashboard'))
    

@app.route('/credit_card_statement/<int:card_id>', methods=['GET'])
def credit_card_statement(card_id):
    """View credit card statement and transactions"""
    if not is_session_valid():
        return redirect(url_for('login'))
    
    try:
        credit_card = CreditCard.query.get(card_id)
        account = Account.query.filter_by(acno=session['user_acno']).first()
        
        if not credit_card or credit_card.acno != account.acno:
            flash('Credit card not found or unauthorized.', 'error')
            return redirect(url_for('my_credit_cards'))
        
        transactions = CreditCardTransaction.query.filter_by(card_id=credit_card.id).order_by(
            CreditCardTransaction.transaction_date.desc()
        ).all()
        
        return render_template('credit_card_statement.html', 
                             credit_card=credit_card,
                             transactions=transactions,
                             csrf_token=generate_csrf())
    except Exception as e:
        app.logger.error(f"Error loading credit card statement: {str(e)}")
        flash('An error occurred while loading credit card statement.', 'error')
        return redirect(url_for('my_credit_cards'))


@app.route('/credit_card_payment/<int:card_id>', methods=['GET', 'POST'])
def credit_card_payment(card_id):
    """Make credit card payment"""
    if not is_session_valid():
        return redirect(url_for('login'))
    
    try:
        credit_card = CreditCard.query.get(card_id)
        account = Account.query.filter_by(acno=session['user_acno']).first()
        
        if not credit_card or credit_card.acno != account.acno:
            flash('Credit card not found or unauthorized.', 'error')
            return redirect(url_for('my_credit_cards'))
        
        amount_owed = credit_card.current_balance
        
        if amount_owed <= 0:
            flash('✅ No outstanding balance. Your credit card account is in good standing!', 'info')
            return redirect(url_for('my_credit_cards'))
        
        if request.method == 'POST':
            try:
                payment_amount = int(request.form.get('payment_amount', 0))
                user_amount = Amount.query.filter_by(acno=account.acno).first()
                
                if payment_amount <= 0:
                    flash('Payment amount must be greater than 0.', 'error')
                    return render_template('credit_card_payment.html', 
                                         credit_card=credit_card, 
                                         amount_owed=amount_owed,
                                         csrf_token=generate_csrf())
                
                if payment_amount > amount_owed:
                    flash(f'Payment amount cannot exceed amount owed (₹{amount_owed:,}). You cannot overpay.', 'error')
                    return render_template('credit_card_payment.html', 
                                         credit_card=credit_card, 
                                         amount_owed=amount_owed,
                                         csrf_token=generate_csrf())
                
                if user_amount.balance < payment_amount:
                    flash(f'Insufficient balance in your account. Your balance: ₹{user_amount.balance:,}', 'error')
                    return render_template('credit_card_payment.html', 
                                         credit_card=credit_card, 
                                         amount_owed=amount_owed,
                                         csrf_token=generate_csrf())
                
                user_amount.balance -= payment_amount
                credit_card.current_balance -= payment_amount
                credit_card.available_credit += payment_amount
                
                if credit_card.available_credit > credit_card.credit_limit:
                    credit_card.available_credit = credit_card.credit_limit
                
                amount_owed_after = credit_card.current_balance
                
                account_txn = Transaction(
                    acno=account.acno,
                    transaction_type='CREDIT_CARD_PAYMENT',
                    amount=payment_amount,
                    balance_after=user_amount.balance,
                    timestamp=datetime.now(),
                    description=f'Credit card payment - Card ending {credit_card.card_number[-4:]}'
                )
                
                cc_txn = CreditCardTransaction(
                    card_id=credit_card.id,
                    transaction_type='PAYMENT',
                    amount=payment_amount,
                    merchant='YourBank - Account Payment',
                    transaction_date=datetime.now(),
                    status='COMPLETED',
                    description=f'Payment received from savings account'
                )
                
                db.session.add(account_txn)
                db.session.add(cc_txn)
                db.session.commit()
                
                send_credit_card_payment_email(account, credit_card, payment_amount)
                
                flash(f'✅ Payment of ₹{payment_amount:,} successful!<br>Amount owed: ₹{amount_owed_after:,} | Available credit: ₹{credit_card.available_credit:,}', 'success')
                return redirect(url_for('my_credit_cards'))
            
            except ValueError:
                flash('Please enter a valid payment amount.', 'error')
                return render_template('credit_card_payment.html', 
                                     credit_card=credit_card, 
                                     amount_owed=amount_owed,
                                     csrf_token=generate_csrf())
        
        return render_template('credit_card_payment.html', 
                             credit_card=credit_card, 
                             amount_owed=amount_owed,
                             csrf_token=generate_csrf())
    except Exception as e:
        app.logger.error(f"Error processing credit card payment: {str(e)}")
        flash('An error occurred while processing payment.', 'error')
        return redirect(url_for('my_credit_cards'))



@app.route('/credit_card_purchase/<int:card_id>', methods=['GET', 'POST'])
def credit_card_purchase(card_id):
    """Make a purchase using credit card"""
    if not is_session_valid():
        return redirect(url_for('login'))
    
    try:
        credit_card = CreditCard.query.get(card_id)
        account = Account.query.filter_by(acno=session['user_acno']).first()
        
        if not credit_card or credit_card.acno != account.acno:
            flash('Credit card not found or unauthorized.', 'error')
            return redirect(url_for('my_credit_cards'))
        
        if request.method == 'POST':
            try:
                merchant_name = request.form.get('merchant_name', 'Online Purchase').strip()
                purchase_amount = int(request.form.get('purchase_amount', 0))
                
                if purchase_amount <= 0:
                    flash('Purchase amount must be greater than 0.', 'error')
                    return render_template('credit_card_purchase.html', 
                                         credit_card=credit_card,
                                         csrf_token=generate_csrf())
                
                if purchase_amount > credit_card.available_credit:
                    flash(f'Purchase amount exceeds available credit (₹{credit_card.available_credit:,}).', 'error')
                    return render_template('credit_card_purchase.html', 
                                         credit_card=credit_card,
                                         csrf_token=generate_csrf())
                
                credit_card.current_balance += purchase_amount
                
                credit_card.available_credit -= purchase_amount
                
                cc_txn = CreditCardTransaction(
                    card_id=credit_card.id,
                    transaction_type='PURCHASE',
                    amount=purchase_amount,
                    merchant=merchant_name,
                    transaction_date=datetime.now(),
                    status='COMPLETED',
                    description=f'Purchase at {merchant_name}'
                )
                
                db.session.add(cc_txn)
                db.session.commit()
                
                send_credit_card_transaction_email(account, cc_txn)
                
                flash(f'✅ Purchase of ₹{purchase_amount:,} successful at {merchant_name}!<br>Amount owed: ₹{credit_card.current_balance:,} | Available credit: ₹{credit_card.available_credit:,}', 'success')
                return redirect(url_for('my_credit_cards'))
            
            except ValueError:
                flash('Please enter a valid purchase amount.', 'error')
                return render_template('credit_card_purchase.html', 
                                     credit_card=credit_card,
                                     csrf_token=generate_csrf())
        
        return render_template('credit_card_purchase.html', 
                             credit_card=credit_card,
                             csrf_token=generate_csrf())
    except Exception as e:
        app.logger.error(f"Error processing credit card purchase: {str(e)}")
        flash('An error occurred while processing purchase.', 'error')
        return redirect(url_for('my_credit_cards'))
        
@app.route('/bulk_payments_upload', methods=['POST'])
def bulk_payments_upload():
    """Process bulk payments upload"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    try:
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('bulk_payments'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('bulk_payments'))
        
        account = Account.query.filter_by(acno=session['user_acno']).first()
        user_amount = Amount.query.filter_by(acno=account.acno).first()
        
        payments = []
        total_amount = 0
        
        if file.filename.endswith('.csv'):
            stream = StringIO(file.stream.read().decode('UTF8'), newline=None)
            csv_data = csv.DictReader(stream)
            
            for row in csv_data:
                try:
                    acno = row.get('Account Number', '').strip()
                    amount = int(row.get('Amount', 0))
                    description = row.get('Description', '')
                    
                    if not acno or amount <= 0:
                        continue
                    
                    payments.append({
                        'acno': acno,
                        'amount': amount,
                        'description': description
                    })
                    total_amount += amount
                except Exception:
                    continue
        
        elif file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    acno = str(row[0]).strip() if row[0] else ''
                    amount = int(row[1]) if row[1] else 0
                    description = str(row[2]) if row[2] else ''
                    
                    if not acno or amount <= 0:
                        continue
                    
                    payments.append({
                        'acno': acno,
                        'amount': amount,
                        'description': description
                    })
                    total_amount += amount
                except Exception:
                    continue
        
        if not payments:
            flash('No valid payments found in the file.', 'error')
            return redirect(url_for('bulk_payments'))
        
        if user_amount.balance < total_amount:
            flash(f'Insufficient balance. Required: ₹{total_amount:,}, Available: ₹{user_amount.balance:,}', 'error')
            return redirect(url_for('bulk_payments'))
        
        successful_payments = 0
        failed_payments = 0
        
        for payment in payments:
            try:
                receiver_amount = Amount.query.filter_by(acno=payment['acno']).first()
                if not receiver_amount:
                    failed_payments += 1
                    continue
                
                user_amount.balance -= payment['amount']
                
                receiver_amount.balance += payment['amount']
                
                sender_txn = Transaction(
                    acno=account.acno,
                    transaction_type='BULK_TRANSFER',
                    amount=payment['amount'],
                    balance_after=user_amount.balance,
                    timestamp=datetime.now(),
                    description=payment.get('description', f'Bulk transfer to {payment["acno"]}')
                )
                
                receiver_txn = Transaction(
                    acno=payment['acno'],
                    transaction_type='RECEIVE',
                    amount=payment['amount'],
                    balance_after=receiver_amount.balance,
                    timestamp=datetime.now(),
                    description=f'Bulk transfer from {account.acno}'
                )
                
                db.session.add(sender_txn)
                db.session.add(receiver_txn)
                successful_payments += 1
            except Exception as e:
                app.logger.error(f"Error processing bulk payment: {str(e)}")
                failed_payments += 1
                continue
        
        db.session.commit()
        flash(f'Bulk payments processed! Successful: {successful_payments}, Failed: {failed_payments}', 'success')
        return redirect(url_for('bulk_payments'))
        
    except Exception as e:
        app.logger.error(f"Error uploading bulk payments: {str(e)}")
        flash('An error occurred while processing bulk payments.', 'error')
        return redirect(url_for('bulk_payments'))


@app.route('/invest', methods=['GET', 'POST'])
def invest():
    """Invest Now - Create new investment"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            investment_type = request.form.get('investment_type', 'MUTUAL_FUND')
            amount = int(request.form.get('amount', 0))
            
            if amount < 5000:
                flash('Minimum investment amount is ₹5,000.', 'error')
                return render_template('invest.html', csrf_token=generate_csrf())
            
            account = Account.query.filter_by(acno=session['user_acno']).first()
            if not account:
                flash('Account not found.', 'error')
                return redirect(url_for('login'))
            
            amount_record = Amount.query.filter_by(acno=session['user_acno']).first()
            
            if not amount_record or amount_record.balance < amount:
                flash('Insufficient balance. Please check your account balance.', 'error')
                return render_template('invest.html', csrf_token=generate_csrf())
            
            amount_record.balance -= amount
            
            try:
                investment_manager = InvestmentManager()
                purchase_price = investment_manager.get_current_price(investment_type)
                annual_return = investment_manager.get_expected_return(investment_type)
            except Exception as e:
                app.logger.error(f"Error with InvestmentManager: {e}")
                purchase_price = 100
                annual_return = 8
            
            quantity = amount / purchase_price if purchase_price > 0 else amount / 100
            
            investment = Investment(
                acno=account.acno,
                investment_id=f"INV{random.randint(100000, 999999)}",
                investment_type=investment_type,
                amount=amount,
                quantity=quantity,
                purchase_price=purchase_price,
                current_price=purchase_price,
                annual_return=annual_return,
                invested_at=datetime.now(),
                status='ACTIVE'
            )
            
            txn = Transaction(
                acno=account.acno,
                transaction_type='INVESTMENT',
                amount=amount,
                balance_after=amount_record.balance,
                timestamp=datetime.now(),
                description=f'Investment in {investment_type} - {quantity:.2f} units @ ₹{purchase_price:.2f}'
            )
            
            db.session.add(investment)
            db.session.add(txn)
            db.session.commit()
            
            send_investment_confirmation_email(account, investment)
            flash(f'Investment created successfully! Investment ID: {investment.investment_id}', 'success')
        
        except Exception as e:
            app.logger.error(f"Error creating investment: {str(e)}")
            flash('An error occurred while creating investment.', 'error')
    
    return render_template('invest.html', csrf_token=generate_csrf())


@app.route('/setup_2fa', methods=['GET', 'POST'])
def setup_2fa():
    """Setup Two-Factor Authentication"""
    if not is_session_valid():
        flash('Your session has expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    user_acno = session.get('user_acno')
    account = Account.query.filter_by(acno=user_acno).first()
    
    if request.method == 'POST':
        try:
            action = request.form.get('action')
            
            if action == 'enable':
                secret = pyotp.random_base32()
                account.two_factor_secret = secret
                account.two_factor_enabled = False  
                db.session.commit()
                
                totp = pyotp.TOTP(secret)
                qr = qrcode.QRCode(version=1, box_size=10, border=5)
                qr.add_data(totp.provisioning_uri(name=account.email, issuer_name='YourBank'))
                qr.make(fit=True)
                
                img = qr.make_image(fill_color="black", back_color="white")
                img_io = BytesIO()
                img.save(img_io, 'PNG')
                img_io.seek(0)
                qr_code_base64 = base64.b64encode(img_io.getvalue()).decode()
                
                flash('2FA setup initiated. Scan the QR code with your authenticator app.', 'info')
                return render_template('setup_2fa.html', qr_code=qr_code_base64, 
                                     secret=secret, step='verify', csrf_token=generate_csrf())
            
            elif action == 'verify':
                otp_code = request.form.get('otp_code', '').strip()
                secret = request.form.get('secret', '').strip()
                
                if not otp_code or len(otp_code) != 6:
                    flash('Please enter a valid 6-digit OTP code.', 'error')
                    return render_template('setup_2fa.html', secret=secret, 
                                         step='verify', csrf_token=generate_csrf())
                
                totp = pyotp.TOTP(secret)
                if totp.verify(otp_code):
                    account.two_factor_enabled = True
                    db.session.commit()
                    flash('Two-Factor Authentication enabled successfully!', 'success')
                    return redirect(url_for('dashboard'))
                else:
                    flash('Invalid OTP code. Please try again.', 'error')
                    return render_template('setup_2fa.html', secret=secret, 
                                         step='verify', csrf_token=generate_csrf())
            
            elif action == 'disable':
                password = request.form.get('password', '')
                if not check_password_hash(account.password, password):
                    flash('Invalid password.', 'error')
                    return render_template('setup_2fa.html', step='disable', csrf_token=generate_csrf())
                
                account.two_factor_enabled = False
                account.two_factor_secret = None
                db.session.commit()
                flash('Two-Factor Authentication disabled.', 'success')
                return redirect(url_for('dashboard'))
        
        except Exception as e:
            app.logger.error(f"Error setting up 2FA: {str(e)}")
            flash('An error occurred. Please try again.', 'error')
    
    return render_template('setup_2fa.html', step='initial', csrf_token=generate_csrf(), 
                         two_factor_enabled=account.two_factor_enabled)

@app.route('/verify_2fa', methods=['POST', 'GET'])
def verify_2fa():
    """Verify 2FA code during login"""
    if request.method == 'GET':
        return render_template('verify_2fa.html', csrf_token=generate_csrf())
    
    acno = session.get('temp_acno')
    if not acno:
        flash('Session expired. Please login again.', 'error')
        return redirect(url_for('login'))
    
    otp_code = request.form.get('otp_code', '').strip()
    
    if not otp_code or len(otp_code) != 6:
        flash('Please enter a valid 6-digit OTP code.', 'error')
        return render_template('verify_2fa.html', csrf_token=generate_csrf())
    
    try:
        account = Account.query.filter_by(acno=acno).first()
        
        if not account or not account.two_factor_secret:
            flash('2FA not configured.', 'error')
            return redirect(url_for('login'))
        
        totp = pyotp.TOTP(account.two_factor_secret)
        if totp.verify(otp_code, valid_window=1):  
            session['user_acno'] = account.acno
            session['user_name'] = account.name
            session['account_type'] = account.account_type
            session['last_activity'] = datetime.now().timestamp()
            session.pop('temp_acno', None)
            
            flash('Login successful!', 'success')
            if account.account_type == 'CURRENT':
                return redirect(url_for('current_account_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        else:
            flash('Invalid OTP code. Please try again.', 'error')
            return render_template('verify_2fa.html', csrf_token=generate_csrf())
    
    except Exception as e:
        app.logger.error(f"Error verifying 2FA: {str(e)}")
        flash('An error occurred during verification.', 'error')
        return redirect(url_for('login'))


def process_emi_deductions():
    """Process monthly EMI deductions for all active loans"""
    with app.app_context():
        try:
            today = datetime.now().date()
            active_loans = LoanAccount.query.filter_by(status='ACTIVE').all()
            
            for loan in active_loans:
                if loan.next_emi_date and loan.next_emi_date.date() == today:
                    account = Account.query.filter_by(acno=loan.acno).first()
                    amount_record = Amount.query.filter_by(acno=loan.acno).first()
                    
                    if amount_record.balance >= loan.monthly_emi:
                        amount_record.balance -= loan.monthly_emi
                        loan.outstanding_balance -= loan.monthly_emi
                        loan.total_paid += loan.monthly_emi
                        
                        loan.next_emi_date += timedelta(days=30)
                        
                        txn = Transaction(
                            acno=loan.acno,
                            transaction_type='EMI_PAYMENT',
                            amount=loan.monthly_emi,
                            balance_after=amount_record.balance,
                            timestamp=datetime.now()
                        )
                        db.session.add(txn)
                        
                        if loan.outstanding_balance <= 0:
                            loan.status = 'CLOSED'
                            send_loan_closure_email(account, loan)
                        else:
                            send_emi_payment_email(account, loan)
                        
                        db.session.commit()
                        logger.info(f"EMI processed for loan {loan.loan_id}")
                    else:
                        send_emi_failure_email(account, loan)
                        logger.warning(f"EMI failed for loan {loan.loan_id} - insufficient balance")
        except Exception as e:
            logger.error(f"Error processing EMI deductions: {str(e)}")

def credit_fd_interest():
    """Credit interest for matured Fixed Deposits"""
    with app.app_context():
        try:
            today = datetime.now().date()
            matured_fds = FixedDeposit.query.filter_by(status='ACTIVE').filter(
                FixedDeposit.maturity_date <= today
            ).all()
            
            for fd in matured_fds:
                account = Account.query.filter_by(acno=fd.acno).first()
                amount_record = Amount.query.filter_by(acno=fd.acno).first()
                
                interest = fd.maturity_amount - fd.principal_amount
                
                amount_record.balance += fd.maturity_amount
                fd.status = 'MATURED'
                
                txn = Transaction(
                    acno=fd.acno,
                    transaction_type='FD_MATURITY',
                    amount=fd.maturity_amount,
                    balance_after=amount_record.balance,
                    timestamp=datetime.now(),
                    description=f'FD Maturity - {fd.fd_id}'
                )
                db.session.add(txn)
                db.session.commit()
                
                send_fd_maturity_email(account, fd, interest)
                logger.info(f"FD {fd.fd_id} matured and credited")
        except Exception as e:
            logger.error(f"Error crediting FD interest: {str(e)}")

def credit_rd_interest():
    """Process Recurring Deposit payments and credit interest on maturity"""
    with app.app_context():
        try:
            today = datetime.now().date()
            
            active_rds = RecurringDeposit.query.filter_by(status='ACTIVE').all()
            for rd in active_rds:
                if rd.next_payment_date and rd.next_payment_date.date() == today:
                    account = Account.query.filter_by(acno=rd.acno).first()
                    amount_record = Amount.query.filter_by(acno=rd.acno).first()
                    
                    if amount_record.balance >= rd.monthly_amount:
                        amount_record.balance -= rd.monthly_amount
                        rd.total_paid += rd.monthly_amount
                        rd.next_payment_date += timedelta(days=30)
                        
                        txn = Transaction(
                            acno=rd.acno,
                            transaction_type='RD_PAYMENT',
                            amount=rd.monthly_amount,
                            balance_after=amount_record.balance,
                            timestamp=datetime.now()
                        )
                        db.session.add(txn)
                        db.session.commit()
            
            matured_rds = RecurringDeposit.query.filter_by(status='ACTIVE').filter(
                RecurringDeposit.maturity_date <= today
            ).all()
            
            for rd in matured_rds:
                account = Account.query.filter_by(acno=rd.acno).first()
                amount_record = Amount.query.filter_by(acno=rd.acno).first()
                
                interest = rd.maturity_amount - rd.total_paid
                amount_record.balance += rd.maturity_amount
                rd.status = 'MATURED'
                
                txn = Transaction(
                    acno=rd.acno,
                    transaction_type='RD_MATURITY',
                    amount=rd.maturity_amount,
                    balance_after=amount_record.balance,
                    timestamp=datetime.now(),
                    description=f'RD Maturity - {rd.rd_id}'
                )
                db.session.add(txn)
                db.session.commit()
                
                send_rd_maturity_email(account, rd, interest)
                logger.info(f"RD {rd.rd_id} matured and credited")
        except Exception as e:
            logger.error(f"Error processing RD: {str(e)}")

def credit_savings_interest():
    """Credit monthly interest to savings accounts"""
    with app.app_context():
        try:
            today = datetime.now()
            last_day = calendar.monthrange(today.year, today.month)[1]
            
            if today.day == last_day:
                savings_accounts = Account.query.filter_by(account_type='SAVINGS').all()
                
                for account in savings_accounts:
                    amount_record = Amount.query.filter_by(acno=account.acno).first()
                    
                    if amount_record and amount_record.balance > 0:
                        interest = (amount_record.balance * 0.04) / 12
                        amount_record.balance += interest
                        
                        txn = Transaction(
                            acno=account.acno,
                            transaction_type='INTEREST_CREDIT',
                            amount=interest,
                            balance_after=amount_record.balance,
                            timestamp=datetime.now(),
                            description='Monthly savings interest'
                        )
                        db.session.add(txn)
                        db.session.commit()
                        
                        send_interest_credit_email(account, interest)
                        logger.info(f"Interest credited to account {account.acno}")
        except Exception as e:
            logger.error(f"Error crediting savings interest: {str(e)}")
try:
    if not scheduler.running:
        scheduler.add_job(process_emi_deductions, 'cron', hour=0, minute=0)
        scheduler.add_job(credit_fd_interest, 'cron', hour=1, minute=0)
        scheduler.add_job(credit_rd_interest, 'cron', hour=2, minute=0)
        scheduler.add_job(credit_savings_interest, 'cron', hour=3, minute=0)
        scheduler.start()
        logger.info("Background schedulers started successfully")
except Exception as e:
    logger.error(f"Error starting scheduler: {str(e)}")

# ==================== MAIN ENTRY POINT ====================
if __name__ == '__main__':
    with app.app_context():
        try:
            db.create_all()
            print("✅ Database tables created successfully")
        except Exception as e:
            print(f"❌ Error creating database tables: {str(e)}")
    
    print("\n" + "="*60)
    print("🚀 Starting YourBank Banking System")
    print("="*60)
    print("\n📱 Access the application at:")
    print("\n   🌐 http://localhost:5001")
    print("\n   Local: http://127.0.0.1:5001")
    print("\n" + "="*60 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=5001, use_reloader=False)
